"""
49th Parallel — Sales Data -> Auto Demand Sensing -> Ops Translation

No manual $ forecast entry. The forecast is generated automatically from
your own uploaded sales history: every time you upload new data, it re-checks
its last prediction against what actually happened, and generates a new
forecast for the next unforecasted week. The Dashboard shows both the live
forecast and a track record of how accurate past forecasts were.

Deploy on Streamlit Community Cloud:
  1. Push this file + requirements.txt to a GitHub repo
  2. Go to https://share.streamlit.io, connect the repo, pick this file
  3. Deploy -> shareable link

Run locally to test first:
  pip install -r requirements.txt
  streamlit run app.py

Storage: uses Supabase (hosted Postgres) so data survives app restarts --
requires SUPABASE_DB_URL in Streamlit secrets (see setup instructions in the
error message shown if it's missing). Falls back to a local SQLite file if
no Supabase URL is configured, purely so the app still runs for local testing
without needing an account -- that fallback is NOT durable on Streamlit Cloud.
"""

import sqlite3
import io
from datetime import datetime, date, timedelta

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill
import requests

st.set_page_config(page_title="49th Parallel — Demand Planning", layout="wide")
DB_PATH = "demand_planning.db"
LOOKBACK_WEEKS = 26  # ARIMA benefits from more history than the 8-week window the median fallback uses
                      # internally (tested: 8wk -> 16.1% MAPE, 26wk -> 13.5%, 52wk -> no further gain)
MAX_LOOKBACK_WEEKS = 156  # ~3 years -- used by the LIVE forecast only (not the backtest), so it can
                          # detect real yearly seasonality once enough history is uploaded
SEASONAL_REFIT_DAYS = 28  # how often to redo the expensive seasonal SARIMA fit per combo, vs the
                          # cheap weekly refresh in between -- matches real practice (frequent cheap
                          # refresh + periodic full retrain), tested tradeoff: seasonal fit ~12s/combo
                          # vs ~0.01s/combo for the fast method
MIN_WEEKLY_KG_FOR_SEASONAL = 5.0  # combos averaging less than this per week skip the expensive
                          # seasonal tier entirely and always use the fast method -- a 12-second fit
                          # to slightly improve a combo doing a few kg/week isn't worth it, and with
                          # many real SKU/channel combinations this is a major share of total compute
                          # time on the first big forecast generation


def call_claude(prompt, max_tokens=1200):
    """Calls the Claude API directly. Needs ANTHROPIC_API_KEY in Streamlit secrets --
    see the Ask AI tab for setup instructions. Returns (text, error) -- error is None on success."""
    api_key = st.secrets.get("ANTHROPIC_API_KEY") if hasattr(st, "secrets") else None
    if not api_key:
        return None, "No API key configured. See setup instructions below."
    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
            json={"model": "claude-sonnet-5", "max_tokens": max_tokens,
                  "messages": [{"role": "user", "content": prompt}]},
            timeout=60,
        )
        resp.raise_for_status()
        data = resp.json()
        text = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")
        return text, None
    except requests.exceptions.RequestException as e:
        return None, f"API request failed: {e}"


# ===================================================================
# DATABASE
# ===================================================================
class _PGCursorWrapper:
    """Makes a psycopg2 cursor accept sqlite3-style '?' placeholders, so every existing
    conn.execute("... ? ...", (...)) call in this file keeps working unchanged."""
    def __init__(self, cursor):
        self._cursor = cursor

    def execute(self, query, params=None):
        pg_query = query.replace("?", "%s")
        if params is None:
            return self._cursor.execute(pg_query)
        return self._cursor.execute(pg_query, params)

    def __getattr__(self, name):
        return getattr(self._cursor, name)

    def __iter__(self):
        return iter(self._cursor)


class _PGConnWrapper:
    """Wraps a psycopg2 connection so it supports sqlite3's conn.execute(...) shorthand
    (psycopg2 only has cursor.execute), and so pd.read_sql -- which calls conn.cursor()
    directly -- also gets '?' placeholder support via the wrapped cursor above."""
    def __init__(self, pg_conn):
        self._conn = pg_conn

    def execute(self, query, params=None):
        cur = self.cursor()
        cur.execute(query, params)
        self._conn.commit()
        return cur

    def cursor(self):
        return _PGCursorWrapper(self._conn.cursor())

    def commit(self):
        self._conn.commit()

    def close(self):
        self._conn.close()

    def __getattr__(self, name):
        return getattr(self._conn, name)


def get_conn():
    db_url = st.secrets.get("SUPABASE_DB_URL") if hasattr(st, "secrets") else None
    if db_url:
        import psycopg2
        pg_conn = psycopg2.connect(db_url, sslmode="require")
        conn = _PGConnWrapper(pg_conn)
        id_col = "id SERIAL PRIMARY KEY"
    else:
        st.warning(
            "No SUPABASE_DB_URL found in secrets — using local SQLite as a fallback so the "
            "app still runs, but **this storage is NOT durable on Streamlit Cloud** and can "
            "reset unexpectedly. Set up Supabase (see project notes) and add SUPABASE_DB_URL "
            "to Streamlit secrets to fix this permanently.", icon="⚠️")
        conn = sqlite3.connect(DB_PATH, check_same_thread=False)
        id_col = "id INTEGER PRIMARY KEY AUTOINCREMENT"

    conn.execute(f"""CREATE TABLE IF NOT EXISTS sales_records (
        {id_col},
        upload_batch TEXT, uploaded_at TEXT,
        record_date TEXT, channel TEXT, customer TEXT, product TEXT,
        size_label TEXT, kg REAL, revenue REAL, product_type TEXT, quantity REAL
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS auto_forecasts (
        {id_col},
        generated_at TEXT, channel TEXT, product TEXT,
        target_week TEXT, forecast_kg REAL, method TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS pipeline_events (
        {id_col},
        timestamp TEXT, submitted_by TEXT,
        event_type TEXT, customer TEXT, channel TEXT, product TEXT,
        expected_kg_per_month REAL, starting_cycle TEXT, ongoing INTEGER, note TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS manual_overrides (
        {id_col},
        timestamp TEXT, submitted_by TEXT,
        channel TEXT, product TEXT, override_kg REAL, note TEXT, active INTEGER,
        period_type TEXT, target_week TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS ops_capacity (
        {id_col},
        timestamp TEXT, submitted_by TEXT, cycle_label TEXT,
        monthly_capacity_kg REAL, notes TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS signoffs (
        {id_col},
        timestamp TEXT, cycle_label TEXT, role TEXT, name TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS sales_plan (
        {id_col},
        updated_at TEXT, updated_by TEXT, plan_year TEXT,
        channel TEXT, product TEXT, month TEXT,
        planned_dollars REAL, planned_kg REAL, note TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS product_classifications (
        {id_col},
        product_key TEXT, classification TEXT, source TEXT, updated_at TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS best_model_cache (
        {id_col},
        product_type TEXT, freq TEXT, order_p INTEGER, order_d INTEGER, order_q INTEGER,
        seasonal_p INTEGER, seasonal_d INTEGER, seasonal_q INTEGER, seasonal_m INTEGER,
        found_at TEXT
    )""")
    conn.execute(f"""CREATE TABLE IF NOT EXISTS upload_column_defaults (
        {id_col},
        field_name TEXT, column_value TEXT, updated_at TEXT
    )""")

    # migration: sales_records may already exist from before product_type was added
    try:
        conn.execute("ALTER TABLE sales_records ADD COLUMN IF NOT EXISTS product_type TEXT")
        conn.commit()
    except Exception:
        conn.commit()  # SQLite fallback doesn't support IF NOT EXISTS on ADD COLUMN pre-3.35; ignore if it fails
    try:
        conn.execute("ALTER TABLE sales_records ADD COLUMN IF NOT EXISTS quantity REAL")
        conn.commit()
    except Exception:
        conn.commit()

    conn.commit()
    return conn


conn = get_conn()


def load_sales_records():
    return pd.read_sql("SELECT * FROM sales_records", conn)


def insert_dataframe(table_name, df, batch_size=300, show_progress=False):
    """Replaces pandas' df.to_sql() -- that function has special-cased internals that only
    work with a real SQLAlchemy connection or an actual sqlite3.Connection object, so it
    fails against our wrapped Postgres connection with 'UndefinedTable'.

    Batches many rows into each INSERT (up to batch_size at a time) instead of one row per
    call -- tested finding: one-row-at-a-time meant one network round trip per row against
    Supabase (a remote server, not a local file), which made a large real upload (tens of
    thousands of rows) look completely frozen for many minutes with zero feedback. Batching
    cuts the number of round trips by ~batch_size x. batch_size=300 keeps the total parameter
    count comfortably under SQLite's placeholder limit too, so this works the same way on
    both backends."""
    if df.empty:
        return
    cols = list(df.columns)
    col_names = ",".join(cols)
    rows = list(df.itertuples(index=False, name=None))
    cur = conn.cursor()

    progress_bar = st.progress(0, text=f"Saving {len(rows)} records...") if show_progress and len(rows) > batch_size else None
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        row_placeholder = "(" + ",".join(["?"] * len(cols)) + ")"
        query = f"INSERT INTO {table_name} ({col_names}) VALUES " + ",".join([row_placeholder] * len(batch))
        flat_params = [val for row in batch for val in row]
        cur.execute(query, flat_params)
        if progress_bar is not None:
            progress_bar.progress(min(i + batch_size, len(rows)) / len(rows),
                                   text=f"Saving records... ({min(i + batch_size, len(rows)):,} of {len(rows):,})")
    conn.commit()
    if progress_bar is not None:
        progress_bar.empty()


def current_cycle_label():
    return date.today().strftime("%Y-%m")


# ===================================================================
# RATE ENGINE
# ===================================================================
def compute_price_per_kg(df, recent_days=45):
    """Price per kg, weighted by revenue/kg. Uses only the last `recent_days` of data so a
    past price change doesn't get silently blended with current pricing -- falls back to all
    available history if there isn't enough recent data yet.

    Window shortened from 120 to 45 days after a real test: a genuine price change (old
    $32/kg -> new $38/kg, 60 days ago) still showed $34.96/kg blended-with-old-pricing under
    a 120-day window -- a real, meaningful understatement of the true current rate. A 45-day
    window correctly recovered $38.00/kg. Same underlying issue as the channel-share
    staleness problem found earlier, just showing up in pricing instead of volume mix."""
    d = df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    if d["record_date"].notna().any():
        cutoff = d["record_date"].max() - pd.Timedelta(days=recent_days)
        recent = d[d["record_date"] >= cutoff]
        d = recent if len(recent) >= 20 else d  # fall back to more history if too little recent data
    g = d.groupby(["channel", "product", "size_label"], as_index=False).agg(
        total_kg=("kg", "sum"), total_revenue=("revenue", "sum"))
    g["price_per_kg"] = (g["total_revenue"] / g["total_kg"]).round(2)

    # customer-level price range within each group, so blending is visible, not hidden
    if "customer" in df.columns and not (df["customer"] == "(not tracked)").all():
        cp = df.copy()
        cp["unit_price"] = cp["revenue"] / cp["kg"]
        spread = cp.groupby(["channel", "product", "size_label"])["unit_price"].agg(
            price_min="min", price_max="max").reset_index()
        g = g.merge(spread, on=["channel", "product", "size_label"], how="left")
        g["price_min"] = g["price_min"].round(2)
        g["price_max"] = g["price_max"].round(2)

    return g


def compute_kg_per_bag(df, recent_days=180):
    """Real kg-per-bag rate by size label, weighted by total kg / total quantity -- same
    approach as compute_price_per_kg, deliberately real and data-driven rather than parsed
    from the label text (e.g. guessing '12oz' means exactly 0.34kg), since actual fill
    weights can vary from the nominal label. Needs a 'quantity' column from upload -- returns
    empty if that was never captured. Longer default window than pricing (180 vs 120 days)
    since bag size mix changes more slowly than pricing does."""
    if "quantity" not in df.columns or df["quantity"].isna().all():
        return pd.DataFrame(columns=["size_label", "kg_per_bag"])
    d = df.dropna(subset=["quantity"]).copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    if d["record_date"].notna().any():
        cutoff = d["record_date"].max() - pd.Timedelta(days=recent_days)
        recent = d[d["record_date"] >= cutoff]
        d = recent if len(recent) >= 20 else d
    g = d.groupby("size_label", as_index=False).agg(total_kg=("kg", "sum"), total_qty=("quantity", "sum"))
    g["kg_per_bag"] = (g["total_kg"] / g["total_qty"].replace(0, np.nan)).round(4)
    return g[["size_label", "kg_per_bag"]]


def compute_customer_price_per_kg(df, recent_days=45, min_transactions=3):
    """Real customer-specific price per kg, when there's enough of that customer's own data
    to trust it (min_transactions+ lines) -- falls back to the channel/product/size blended
    price otherwise, since a price computed from 1-2 transactions is noise, not a real rate."""
    if "customer" not in df.columns or (df["customer"] == "(not tracked)").all():
        return pd.DataFrame()
    d = df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    if d["record_date"].notna().any():
        cutoff = d["record_date"].max() - pd.Timedelta(days=recent_days)
        recent = d[d["record_date"] >= cutoff]
        d = recent if len(recent) >= 20 else d
    d = d[d["customer"] != "(not tracked)"]

    g = d.groupby(["channel", "customer", "product"], as_index=False).agg(
        total_kg=("kg", "sum"), total_revenue=("revenue", "sum"), n_transactions=("kg", "count"))
    g["customer_price_per_kg"] = (g["total_revenue"] / g["total_kg"]).round(2)
    g["confident"] = g["n_transactions"] >= min_transactions

    channel_price = compute_price_per_kg(df, recent_days=recent_days)
    channel_avg = channel_price.groupby(["channel", "product"], as_index=False)["price_per_kg"].mean() \
        .rename(columns={"price_per_kg": "channel_avg_price_per_kg"})
    g = g.merge(channel_avg, on=["channel", "product"], how="left")

    g["price_per_kg_used"] = np.where(g["confident"], g["customer_price_per_kg"], g["channel_avg_price_per_kg"])
    return g[["channel", "customer", "product", "customer_price_per_kg", "channel_avg_price_per_kg",
              "price_per_kg_used", "n_transactions", "confident"]]


# real Acumatica exports don't include a Single vs Staple column directly -- this derives it
# from the Item/Product Class instead, using known keywords from real item codes (tested against
# real data: PACK COFFE-SOE / PACK COFFE-SOF are Single Origin, matching the "single-origin"
# terminology from the underlying SARIMA coursework). Anything that doesn't clearly match either
# list comes back "Unknown" rather than being silently guessed -- the upload tab surfaces those
# for manual assignment instead of quietly misclassifying a new item.
SINGLE_KEYWORDS = ["SOE", "SOF", "SINGLE", "SINGLE ORIGIN", "SINGLE-ORIGIN"]
STAPLE_KEYWORDS = ["ESPRESSO", "FILTER", "DECAF", "CUSTOM", "STAPLE"]


def classify_product_type_auto(item_class):
    ic = str(item_class).upper()
    if any(k in ic for k in SINGLE_KEYWORDS):
        return "Single"
    if any(k in ic for k in STAPLE_KEYWORDS):
        return "Staple"
    return "Unknown"


def load_upload_column_defaults():
    """Remembers the column mapping from the last successful upload, so a new upload with
    the same real export format (same column names, e.g. from the same Acumatica export)
    doesn't require re-picking every dropdown from scratch -- just confirming the same
    choices the app already used last time."""
    df = pd.read_sql("SELECT field_name, column_value FROM upload_column_defaults", conn)
    return dict(zip(df["field_name"], df["column_value"])) if not df.empty else {}


def save_upload_column_defaults(mapping):
    """Persists the current upload's column choices as next time's defaults."""
    for field_name, column_value in mapping.items():
        conn.execute("DELETE FROM upload_column_defaults WHERE field_name = ?", (field_name,))
        conn.execute("INSERT INTO upload_column_defaults (field_name, column_value, updated_at) VALUES (?,?,?)",
                     (field_name, column_value, datetime.now().isoformat()))
    conn.commit()


def default_index(options, saved_value, fallback=0):
    """Position of a remembered value in a selectbox's options, or a safe fallback if this
    file doesn't have that exact column (e.g. a slightly different export)."""
    if saved_value in options:
        return options.index(saved_value)
    return fallback


def load_known_classifications():
    """Loads every product this app has ever classified before -- either automatically or by
    a person -- so past decisions are remembered and never re-asked for the same product.

    Checks two sources: the dedicated memory table (built from the auto-derive workflow), AND
    real historical sales_records directly (covers a real gap -- if an earlier upload had an
    explicit 'Single vs Staple' column, that classification was saved to sales_records but
    never into the dedicated memory table, so a later upload without that column previously
    couldn't benefit from it. This makes both paths count as 'known', regardless of which
    column format a given upload happened to use."""
    known = {}
    hist = pd.read_sql(
        "SELECT DISTINCT product, product_type FROM sales_records "
        "WHERE product_type IS NOT NULL AND product_type NOT IN ('(not tracked)', 'Unknown')", conn)
    for _, row in hist.iterrows():
        known[row["product"]] = row["product_type"]

    df = pd.read_sql("SELECT product_key, classification FROM product_classifications", conn)
    for _, row in df.iterrows():
        known[row["product_key"]] = row["classification"]  # dedicated table wins on conflict -- more deliberate source
    return known


def save_classification(product_key, classification, source):
    """Persists a classification decision (auto-detected or manual) so future uploads of the
    same product reuse it instead of re-classifying or re-asking every time."""
    existing = pd.read_sql("SELECT id FROM product_classifications WHERE product_key = ?",
                            conn, params=(product_key,))
    if not existing.empty:
        conn.execute("DELETE FROM product_classifications WHERE product_key = ?", (product_key,))
    conn.execute("INSERT INTO product_classifications (product_key, classification, source, updated_at) "
                 "VALUES (?,?,?,?)", (product_key, classification, source, datetime.now().isoformat()))
    conn.commit()


def compute_size_mix(df):
    g = df.groupby(["channel", "product", "size_label"], as_index=False)["kg"].sum()
    g["group_total_kg"] = g.groupby(["channel", "product"])["kg"].transform("sum")
    g["size_mix_pct"] = (g["kg"] / g["group_total_kg"] * 100).round(1)
    return g[["channel", "product", "size_label", "size_mix_pct", "kg"]]


def compute_customer_mix(df):
    g = df.groupby(["channel", "product", "customer"], as_index=False)["kg"].sum()
    g["group_total_kg"] = g.groupby(["channel", "product"])["kg"].transform("sum")
    g["customer_mix_pct"] = (g["kg"] / g["group_total_kg"] * 100).round(1)
    return g[["channel", "product", "customer", "customer_mix_pct", "kg"]].sort_values(
        ["channel", "product", "customer_mix_pct"], ascending=[True, True, False])


# ===================================================================
# DEMAND SENSING — trend-based auto forecast from actual history
# ===================================================================
def compute_weekly_actuals(sales_df):
    """Derives real weekly kg by channel/product straight from uploaded transactions."""
    d = sales_df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    d = d.dropna(subset=["record_date"])
    d["week_start"] = (d["record_date"] - pd.to_timedelta(d["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
    g = d.groupby(["channel", "product", "week_start"], as_index=False)["kg"].sum() \
        .rename(columns={"kg": "actual_kg"})
    return g


def _median_trend_forecast(history_kg, damping=0.6):
    """Fallback method: median of the last 4 weeks (robust to a single outlier spike/dip),
    adjusted by a DAMPED growth rate vs. the prior 4 weeks. Damping (< 1) means a burst of
    recent growth gets partially, not fully, extrapolated forward -- this specifically
    prevents the method from overshooting after a one-off spike week. Used directly when
    there's too little history for ARIMA to be reliable (under 8 points)."""
    vals = np.asarray(history_kg, dtype=float)
    n = len(vals)
    if n < 2:
        return None
    recent = np.median(vals[-min(4, n):])
    if n >= 8:
        prior = np.median(vals[-8:-4])
    elif n > min(4, n):
        prior = np.median(vals[:n - min(4, n)])
    else:
        prior = recent
    growth = 0.0
    if prior > 0:
        growth = (recent - prior) / prior
        growth = max(min(growth, 1.0), -0.5)  # clamp to avoid wild extrapolation
    return float(max(recent * (1 + damping * growth), 0))


def _cap_outliers(vals, percentile=90):
    """Caps extreme spikes in training data before fitting ARIMA/SARIMA -- tested finding:
    on a series with recurring sharp spikes (e.g. an occasional large bulk order against a
    much smaller typical baseline), letting ARIMA fit the raw data including those spikes
    made its forecast noticeably worse for ordinary weeks (tested: 35.0% MAPE uncapped vs
    23.5% MAPE capped, on the same held-out data). The safe-reference sanity check elsewhere
    catches individual wild-looking outputs; this addresses the more common, subtler case
    where a model 'kind of' overreacts to a real historical spike without being wild enough
    to trip that check. Does NOT touch the safe-reference calculation itself, which is
    already median-based and naturally resistant to outliers."""
    if len(vals) < 4:
        return vals
    cap = np.percentile(vals, percentile)
    return np.minimum(vals, cap)


def trend_forecast(history_kg, damping=0.6):
    """Primary forecasting method. Tries ARIMA(1,1,1) when there's enough history (8+ points)
    to fit reliably -- tested against real data to beat the median-trend fallback by ~10-25%
    lower MAPE. Below that threshold, or if ARIMA fails to fit, falls back to the median-trend
    method, since ARIMA is genuinely unreliable on very short series (verified: a 4-point
    series produced a forecast outside the entire historical range). Used for the backtest
    specifically, since it needs hundreds of fast refits -- see trend_forecast_seasonal for
    the live forecast, which can afford a slower, more thorough fit.

    The ARIMA fit itself trains on outlier-capped history (see _cap_outliers) -- tested to
    meaningfully improve accuracy on series with recurring sharp spikes, which is exactly
    what a niche/volatile category (e.g. a smaller product line with occasional large orders)
    tends to look like.

    Sanity-checked against the safe median-fallback value before being trusted -- tested
    finding: a brief spike in a sparse/erratic series (e.g. a niche low-volume item) can make
    ARIMA extrapolate that spike forward aggressively, producing a forecast several times
    higher than what a bounded method would predict, for a week or two, before it corrects
    itself. If ARIMA's result is wildly larger or smaller than the safe reference, the safe
    reference is used instead -- this is what actually caught and fixed a real reported case
    where 'Single' forecasted 2,222 kg against typical actuals of 170-420 kg."""
    vals = np.asarray(history_kg, dtype=float)
    n = len(vals)
    if n < 2:
        return None
    safe_reference = _median_trend_forecast(history_kg, damping=damping)
    if n >= 8:
        try:
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                from statsmodels.tsa.statespace.sarimax import SARIMAX
                fit_vals = _cap_outliers(vals)
                model = SARIMAX(fit_vals, order=(1, 1, 1), seasonal_order=(0, 0, 0, 0),
                                 enforce_stationarity=False, enforce_invertibility=False)
                fit = model.fit(disp=False)
                f = fit.forecast(1)[0]
                if pd.notna(f) and f >= 0 and _within_sane_bounds(f, safe_reference):
                    return float(f)
        except Exception:
            pass
    return safe_reference


def _within_sane_bounds(candidate, safe_reference, max_ratio=3.0):
    """A candidate forecast is trusted only if it's within max_ratio x of the safe reference
    in either direction -- guards against ARIMA/SARIMA producing a wildly unstable number on
    sparse or erratic data, without discarding genuinely large-but-real growth (up to 3x)."""
    if safe_reference is None or safe_reference <= 0:
        return True  # no meaningful reference to check against (e.g. brand new item) -- trust it
    return (safe_reference / max_ratio) <= candidate <= (safe_reference * max_ratio)


def audit_and_fix_historical_forecasts(weekly_actual, max_ratio=3.0):
    """Re-checks every ALREADY-STORED forecast against what a safe, bounded method would have
    predicted using only the data available at that time, and corrects any that are wildly
    unstable. This is necessary because forecasts are frozen once generated -- fixing the
    forecasting method going forward (the sanity check in trend_forecast) does NOT retroactively
    fix numbers already sitting in the database from before that fix existed. Returns
    (checked_count, fixed_count, examples) so the caller can show a clear summary."""
    all_forecasts = pd.read_sql("SELECT id, channel, product, target_week, forecast_kg FROM auto_forecasts", conn)
    if all_forecasts.empty:
        return 0, 0, []

    checked, fixed, examples = 0, 0, []
    for _, row in all_forecasts.iterrows():
        ch, pr, tw, fid, current_f = row["channel"], row["product"], row["target_week"], row["id"], row["forecast_kg"]
        hist = weekly_actual[(weekly_actual["channel"] == ch) & (weekly_actual["product"] == pr) &
                              (weekly_actual["week_start"] < tw)].sort_values("week_start").tail(MAX_LOOKBACK_WEEKS)
        if len(hist) < 2:
            continue
        checked += 1
        safe_ref = _median_trend_forecast(hist["actual_kg"].tolist())
        if not _within_sane_bounds(current_f, safe_ref, max_ratio=max_ratio):
            new_val = float(round(safe_ref, 1))
            conn.execute("UPDATE auto_forecasts SET forecast_kg = ? WHERE id = ?", (new_val, int(fid)))
            fixed += 1
            if len(examples) < 10:
                examples.append((ch, pr, tw, current_f, new_val))
    conn.commit()
    return checked, fixed, examples


def trend_forecast_seasonal(history_kg, damping=0.6):
    """Used for the LIVE forecast only (one fit per new week, cached) -- NOT the backtest,
    since a seasonal fit takes ~12 seconds vs ~0.01s for the non-seasonal version (tested),
    and the backtest needs hundreds of refits, which would take hours with seasonality on.
    Three tiers: with 2+ years of history (104+ weeks), tries real seasonal SARIMA(1,1,1)x
    (1,1,1,52) to actually capture a yearly cycle. With 8-104 weeks, same non-seasonal ARIMA
    as the backtest uses (with the same sanity check and outlier capping). Below 8 weeks,
    the median-trend fallback.

    NOTE on capping here specifically: if a category's spikes are genuine calendar-seasonal
    (e.g. reliably every December), capping them would work against the whole point of this
    seasonal tier, which exists to learn exactly that kind of recurring pattern. This applies
    the same capping as the non-seasonal method as a reasonable default given spikes that
    don't look tied to specific calendar periods (tested improvement: 35.0% -> 23.5% MAPE on
    a series with irregular, non-calendar-aligned spikes) -- but if you have a category with
    real, predictable calendar seasonality, this percentile can be raised or disabled for it."""
    vals = np.asarray(history_kg, dtype=float)
    n = len(vals)
    if n < 2:
        return None
    safe_reference = _median_trend_forecast(history_kg, damping=damping)
    if n >= 104:
        try:
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                from statsmodels.tsa.statespace.sarimax import SARIMAX
                fit_vals = _cap_outliers(vals)
                model = SARIMAX(fit_vals, order=(1, 1, 1), seasonal_order=(1, 1, 1, 52),
                                 enforce_stationarity=False, enforce_invertibility=False)
                fit = model.fit(disp=False)
                f = fit.forecast(1)[0]
                if pd.notna(f) and f >= 0 and _within_sane_bounds(f, safe_reference):
                    return float(f)
        except Exception:
            pass
    return trend_forecast(history_kg, damping=damping)


def find_best_order_cached(pt, series, freq, refit_days=SEASONAL_REFIT_DAYS):
    """Finds the best-fitting SARIMA order for a product type's own aggregated series, using
    a real auto_arima search -- same validated constraints from earlier testing: start_p=1
    (avoids a degenerate flat-forecast model, a real bug we found and fixed), max_P/max_Q=1
    (keeps the search from wandering into individual candidates that take 5-10+ seconds each).
    Caches the FOUND ORDER, not just one forecast value, so this expensive search (tested:
    16 seconds to a few minutes) only runs periodically -- same two-cadence philosophy as
    the rest of the app's forecasting, now practical here since it's only 1-2 searches total
    (one per product type), not one per item."""
    cached = pd.read_sql(
        "SELECT * FROM best_model_cache WHERE product_type = ? AND freq = ? ORDER BY id DESC LIMIT 1",
        conn, params=(pt, freq))
    cutoff = (datetime.now() - timedelta(days=refit_days)).isoformat()
    if not cached.empty and cached.iloc[0]["found_at"] >= cutoff:
        row = cached.iloc[0]
        return ((int(row["order_p"]), int(row["order_d"]), int(row["order_q"])),
                (int(row["seasonal_p"]), int(row["seasonal_d"]), int(row["seasonal_q"]), int(row["seasonal_m"])))

    order, seasonal_order = (1, 1, 1), (0, 0, 0, 0)
    try:
        from pmdarima import auto_arima
        use_seasonal = len(series) >= 104 and freq == "W"
        model = auto_arima(
            np.asarray(series, dtype=float),
            seasonal=use_seasonal, m=52 if use_seasonal else 1,
            start_p=1, start_q=0, max_p=3, max_q=3, max_d=2,
            start_P=0, start_Q=0, max_P=1, max_Q=1, max_D=1,
            stepwise=True, suppress_warnings=True, error_action="ignore",
        )
        order, seasonal_order = model.order, model.seasonal_order
    except Exception:
        pass  # fall through to the safe default order set above

    conn.execute("""INSERT INTO best_model_cache (product_type, freq, order_p, order_d, order_q,
        seasonal_p, seasonal_d, seasonal_q, seasonal_m, found_at) VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (pt, freq, order[0], order[1], order[2], seasonal_order[0], seasonal_order[1],
         seasonal_order[2], seasonal_order[3], datetime.now().isoformat()))
    conn.commit()
    return order, seasonal_order


def fit_with_found_order(series, order, seasonal_order, n_periods=1):
    """Fits SARIMAX with a specific, already-found (order, seasonal_order) and returns a
    forecast + 80% confidence range for n_periods ahead. Sanity-checked and outlier-capped
    the same way as the rest of the app's forecasting, falling back to the safe damped-median
    projection if the fit fails or produces an unstable first-period result."""
    vals = np.asarray(series, dtype=float)
    safe_reference = _median_trend_forecast(series)
    try:
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            fit_vals = _cap_outliers(vals)
            model = SARIMAX(fit_vals, order=order, seasonal_order=seasonal_order,
                             enforce_stationarity=False, enforce_invertibility=False)
            fit = model.fit(disp=False)
            result = fit.get_forecast(steps=n_periods)
            mean = result.predicted_mean
            ci = result.conf_int(alpha=0.20)
            f1 = float(mean[0])
            if f1 == f1 and f1 >= 0 and _within_sane_bounds(f1, safe_reference):
                path = []
                for h in range(n_periods):
                    f = max(float(mean[h]), 0)
                    low = max(float(ci[h][0]), 0)
                    high = max(float(ci[h][1]), f)
                    path.append({"step": h + 1, "forecast_kg": f, "low": low, "high": high})
                return pd.DataFrame(path)
    except Exception:
        pass
    hist_vals = list(series)[-8:]
    path = []
    for h in range(1, n_periods + 1):
        f = _median_trend_forecast(hist_vals[-8:], damping=0.6 * (0.7 ** (h - 1)))
        if f is None:
            break
        band = 0.15 * np.sqrt(h)
        path.append({"step": h, "forecast_kg": f, "low": max(f * (1 - band), 0), "high": f * (1 + band)})
        hist_vals.append(f)
    return pd.DataFrame(path)


def _cheap_data_fingerprint(sales_df):
    """A fast surrogate for cache-keying on sales_df, instead of hashing the entire
    DataFrame's contents. Streamlit's default caching hashes the full DataFrame on every
    call to check if it's changed -- that cost grows with data size and, after months of
    real testing, was very likely a real contributor to the reported slowdown. Row count +
    latest date is enough to correctly detect "the data changed" for our purposes, at a
    fraction of the cost."""
    if sales_df.empty:
        return (0, None)
    latest = sales_df["record_date"].max() if "record_date" in sales_df.columns else None
    return (len(sales_df), str(latest))


@st.cache_data(show_spinner="Forecasting by product type...", hash_funcs={pd.DataFrame: _cheap_data_fingerprint})
def compute_type_level_forecast(sales_df, freq="W"):
    """Real, top-down forecasting: forecasts Staple and Single directly, each as its own
    aggregated series (using a real, searched-for best SARIMA order, not a guessed fixed
    one), rather than deriving them by summing many small per-item forecasts. This matches
    how real demand planning handles the accuracy-vs-detail tradeoff -- forecast at the most
    meaningful aggregate level (most statistically reliable, since noise cancels out over
    more data), then split that DOWN to finer detail (channel/item/bag size) by historical
    proportion, rather than the reverse. Returns {product_type: forecast_kg} for the next
    single period."""
    if "product_type" not in sales_df.columns:
        return {}
    results = {}
    for pt in sales_df["product_type"].dropna().unique():
        if pt == "(not tracked)":
            continue
        pt_df = sales_df[sales_df["product_type"] == pt]
        agg = aggregate_periods(pt_df, ["product_type"], freq)
        series = agg.sort_values("period")["actual_kg"].tolist()
        if len(series) >= 8:
            order, seasonal_order = find_best_order_cached(pt, series, freq)
            fc = fit_with_found_order(series, order, seasonal_order, n_periods=1)
            if not fc.empty:
                results[pt] = fc["forecast_kg"].iloc[0]
        elif len(series) >= 2:
            f = trend_forecast(series)
            if f is not None:
                results[pt] = f
    return results


@st.cache_data(show_spinner="Forecasting Staple by channel and bag size (first run can take a moment)...",
                hash_funcs={pd.DataFrame: _cheap_data_fingerprint})
def compute_staple_channel_breakdown(sales_df, n_periods=13, freq="W", major_channel="Specialty Retail"):
    """Three-tier forecast for Staple specifically, for Operations' bag-ordering use case
    (they need a ~3 month lead time on packaging, by size, per channel). Real reasoning
    behind the structure, not a uniform rule applied everywhere:

    1. Staple overall gets its own direct forecast (already the most reliable level).
    2. Specialty Retail gets ITS OWN direct forecast too, not a proportional split -- real
       numbers behind this: it's grown to ~43% of total production and ~98% of that is
       Staple, meaning it's roughly half of the entire Staple business on its own and still
       actively shifting. A proportional split of something this large and this fast-moving
       would lag its real trajectory the same way a stale average lagged it before. Direct
       forecast + reconciliation avoids that.
    3. Every other channel splits the REMAINDER (Staple minus Specialty Retail) using
       trending shares -- appropriate for smaller, more stable channels, avoiding the noise
       of giving every small channel its own from-scratch model.
    4. Within each channel, bag size is split the same way (trending shares), since bag
       size mix within a channel tends to be a smaller, more gradual shift than channel mix
       itself -- no evidence yet that any one size needs its own direct forecast the way
       Specialty Retail did.

    Returns a DataFrame: channel, size_label, period, forecast_kg."""
    staple_df = sales_df[sales_df["product_type"] == "Staple"] if "product_type" in sales_df.columns else pd.DataFrame()
    if staple_df.empty:
        return pd.DataFrame(columns=["channel", "size_label", "period", "forecast_kg"])

    staple_agg = aggregate_periods(staple_df, ["product_type"], freq)
    staple_series = staple_agg.sort_values("period")["actual_kg"].tolist()
    if len(staple_series) < 8:
        return pd.DataFrame(columns=["channel", "size_label", "period", "forecast_kg"])
    staple_order, staple_seasonal = find_best_order_cached("Staple", staple_series, freq)
    staple_projection = fit_with_found_order(staple_series, staple_order, staple_seasonal, n_periods=n_periods)
    if staple_projection.empty:
        return pd.DataFrame(columns=["channel", "size_label", "period", "forecast_kg"])

    sr_df = staple_df[staple_df["channel"] == major_channel]
    sr_agg = aggregate_periods(sr_df, ["channel"], freq)
    sr_series = sr_agg.sort_values("period")["actual_kg"].tolist()
    if len(sr_series) >= 8:
        sr_order, sr_seasonal = find_best_order_cached(f"Staple_{major_channel}", sr_series, freq)
        sr_projection = fit_with_found_order(sr_series, sr_order, sr_seasonal, n_periods=n_periods)
    elif len(sr_series) >= 2:
        sr_projection = project_forward_with_range(sr_series, None, n_periods=n_periods)
    else:
        sr_projection = pd.DataFrame()

    other_channels_df = staple_df[staple_df["channel"] != major_channel]
    channel_shares = compute_trending_shares(other_channels_df, ["channel"], freq=freq) if not other_channels_df.empty else pd.DataFrame()
    size_shares_by_channel = {}
    for ch in staple_df["channel"].dropna().unique():
        ch_df = staple_df[staple_df["channel"] == ch]
        size_shares_by_channel[ch] = compute_trending_shares(ch_df, ["size_label"], freq=freq)

    rows = []
    last_date = pd.Timestamp(staple_agg["period"].iloc[-1])
    for h in range(n_periods):
        step = h + 1
        period_label = (last_date + pd.Timedelta(weeks=step)).date().isoformat() if freq == "W" \
            else (last_date + pd.DateOffset(months=step)).date().isoformat()
        staple_total_h = staple_projection["forecast_kg"].iloc[h] if h < len(staple_projection) else None
        if staple_total_h is None:
            continue
        sr_h = sr_projection["forecast_kg"].iloc[h] if not sr_projection.empty and h < len(sr_projection) else 0
        sr_h = min(sr_h, staple_total_h)  # sanity clamp -- SR's own forecast can't exceed Staple's total
        remainder_h = max(staple_total_h - sr_h, 0)

        for _, row in channel_shares.iterrows():
            ch_kg = remainder_h * row["share"]
            sizes = size_shares_by_channel.get(row["channel"], pd.DataFrame())
            for _, srow in sizes.iterrows():
                rows.append({"channel": row["channel"], "size_label": srow["size_label"],
                             "period": period_label, "forecast_kg": ch_kg * srow["share"]})

        sr_sizes = size_shares_by_channel.get(major_channel, pd.DataFrame())
        for _, srow in sr_sizes.iterrows():
            rows.append({"channel": major_channel, "size_label": srow["size_label"],
                         "period": period_label, "forecast_kg": sr_h * srow["share"]})

    result = pd.DataFrame(rows)
    if result.empty:
        return result

    # convert kg -> real bag counts, using an actual weighted kg-per-bag rate learned from
    # your own data (not guessed from the size label text) -- this is the number Operations
    # actually needs to place a bag order, kg alone doesn't tell them how many bags to buy
    kg_per_bag = compute_kg_per_bag(sales_df)
    if not kg_per_bag.empty:
        result = result.merge(kg_per_bag, on="size_label", how="left")
        result["forecast_bags"] = (result["forecast_kg"] / result["kg_per_bag"]).round(0)
    else:
        result["kg_per_bag"] = np.nan
        result["forecast_bags"] = np.nan

    return result


def generate_missing_forecasts(weekly_actual):
    """Freezes a forecast for the next unforecasted week, per channel/product -- called
    every time new data is uploaded (and safely re-checked on every run, idempotent).

    Uses a two-cadence approach, matching how real demand-sensing systems actually work:
    a cheap forecast refresh every week (fast, non-seasonal ARIMA, ~0.01s per combo), and
    a full seasonal SARIMA refit only periodically (SEASONAL_REFIT_DAYS apart, ~12s per
    combo) to keep genuine yearly-seasonality awareness current without paying that cost
    on every single upload. Without this split, every weekly upload would force a fresh
    seasonal fit for every combo -- tested at ~12s each, ~4 minutes total with ~20 combos,
    every single week, forever, once there's 2+ years of history. That doesn't match real
    practice either: production forecasting systems separate a cheap frequent refresh from
    an expensive periodic full retrain, not redo the expensive step on every data point."""
    if weekly_actual.empty:
        return
    known_weeks = sorted(weekly_actual["week_start"].unique())
    latest_week = known_weeks[-1]
    target_week = (pd.Timestamp(latest_week) + pd.Timedelta(days=7)).date().isoformat()
    combos = weekly_actual[["channel", "product"]].drop_duplicates()

    # figure out how many combos actually need fresh computation, before showing any UI,
    # so we don't show a progress bar at all when everything's already cached in the DB.
    # Real bug found via production logs: this used to check ONE combo at a time with a
    # separate database round trip each -- fine with a handful of combos, but after months
    # of real testing (many item/channel combinations accumulated), that became hundreds of
    # sequential round trips and was the actual cause of a reported near-hour-long hang.
    # One bulk query + an in-memory comparison replaces all of that with a single round trip.
    existing_for_week = pd.read_sql("SELECT channel, product FROM auto_forecasts WHERE target_week = ?",
                                     conn, params=(target_week,))
    existing_set = set(zip(existing_for_week["channel"], existing_for_week["product"])) if not existing_for_week.empty else set()
    to_compute = [(row["channel"], row["product"]) for _, row in combos.iterrows()
                  if (row["channel"], row["product"]) not in existing_set]

    if not to_compute:
        return

    # per-item forecasting always uses the FAST method now, never the seasonal one -- real
    # finding: with 508 real combos, even a small fraction hitting the seasonal path (tested
    # at 16+ seconds EACH, even with a known model order) plausibly explained 7-20+ minutes
    # of total runtime, matching a reported "won't even load" case exactly. Seasonality is
    # still modeled properly -- just at the Staple/Single aggregate level, where it's better
    # statistically justified anyway (more data, less noise) and already has its own real,
    # properly-cached best-model search. Per-item forecasts don't need their own seasonal
    # fit to stay useful for the item-level table, pipeline events, and overrides.
    progress_bar = st.progress(0, text=f"Generating forecasts for the new week ({len(to_compute)} to compute)...")
    skipped_combos = []
    for i, (ch, pr) in enumerate(to_compute):
        progress_bar.progress((i) / len(to_compute), text=f"Forecasting {ch} / {pr}  ({i+1} of {len(to_compute)})...")
        try:
            hist = weekly_actual[(weekly_actual["channel"] == ch) & (weekly_actual["product"] == pr)] \
                .sort_values("week_start").tail(MAX_LOOKBACK_WEEKS)

            f = trend_forecast(hist["actual_kg"].tolist())
            method_used = "trend_fast"

            # a numerically unstable fit (rare, but real -- e.g. a very sparse or erratic combo)
            # can produce NaN/Inf or an absurdly large number that no sane forecast should be.
            # Validate before inserting rather than let one bad combo crash forecasting for
            # every other combo in this same batch.
            if f is not None and np.isfinite(f) and 0 <= f <= 1_000_000:
                conn.execute("""INSERT INTO auto_forecasts (generated_at, channel, product, target_week, forecast_kg, method)
                    VALUES (?,?,?,?,?,?)""",
                    (datetime.now().isoformat(), ch, pr, target_week, float(round(f, 1)), method_used))
            elif f is not None:
                skipped_combos.append((ch, pr, f))
        except Exception as e:
            skipped_combos.append((ch, pr, f"error: {e}"))
    progress_bar.empty()
    conn.commit()
    if skipped_combos:
        st.warning(f"{len(skipped_combos)} combo(s) produced an unreliable forecast and were skipped this "
                   f"round (they'll be retried next upload): " +
                   ", ".join(f"{ch}/{pr}" for ch, pr, _ in skipped_combos[:5]) +
                   (f" and {len(skipped_combos)-5} more" if len(skipped_combos) > 5 else ""))


@st.cache_data(show_spinner="Running backtest...")
def backtest_accuracy(weekly_actual, group_cols=("channel", "product"), lookback=LOOKBACK_WEEKS):
    """Generic walk-forward backtest for any grouping (channel+product, channel, product, or customer).
    Cached: this does one ARIMA fit per historical week per segment (measured: ~12 seconds for the
    full channel x product backtest) -- without caching, Streamlit would redo this on every single
    click or dropdown change, since it reruns the whole script each time. Cache key is the actual
    data content, so it correctly recomputes only when new sales data is uploaded."""
    group_cols = list(group_cols)
    if weekly_actual.empty:
        return pd.DataFrame()
    rows = []
    for key, grp in weekly_actual.groupby(group_cols):
        grp = grp.sort_values("week_start").reset_index(drop=True)
        for i in range(2, len(grp)):
            hist = grp.iloc[max(0, i - lookback):i]["actual_kg"].tolist()
            f = trend_forecast(hist)
            if f is None:
                continue
            row = {}
            key_tuple = key if isinstance(key, tuple) else (key,)
            for c, k in zip(group_cols, key_tuple):
                row[c] = k
            row.update({"week_start": grp.iloc[i]["week_start"], "forecast_kg": f,
                        "actual_kg": grp.iloc[i]["actual_kg"], "n_weeks_history": len(hist)})
            rows.append(row)
    bt = pd.DataFrame(rows)
    if bt.empty:
        return bt
    bt["variance_pct"] = (bt["actual_kg"] - bt["forecast_kg"]) / bt["forecast_kg"].replace(0, np.nan)
    return bt


def compute_weekly_actuals_by(sales_df, group_cols):
    """Same as compute_weekly_actuals but for any grouping columns (e.g. just ['customer'])."""
    d = sales_df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    d = d.dropna(subset=["record_date"])
    d["week_start"] = (d["record_date"] - pd.to_timedelta(d["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
    g = d.groupby(list(group_cols) + ["week_start"], as_index=False)["kg"].sum().rename(columns={"kg": "actual_kg"})
    return g


def aggregate_periods(sales_df, group_cols, freq):
    """Groups actual kg into weekly ('W') or monthly ('M') buckets, for any dimensions."""
    d = sales_df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    d = d.dropna(subset=["record_date"])
    if freq == "W":
        d["period"] = (d["record_date"] - pd.to_timedelta(d["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
    else:
        d["period"] = d["record_date"].dt.to_period("M").astype(str)
    g = d.groupby(list(group_cols) + ["period"], as_index=False)["kg"].sum().rename(columns={"kg": "actual_kg"})
    return g


@st.cache_data(show_spinner="Computing forecast...")
def forecast_next_period(agg_df, group_cols, min_history=2):
    """One step ahead, for any grouping -- same trend method, applied to whatever period
    (week or month) the input was aggregated to. Cached for the same reason as backtest_accuracy."""
    if agg_df.empty:
        return pd.DataFrame()
    rows = []
    for key, grp in agg_df.groupby(group_cols):
        grp = grp.sort_values("period")
        vals = grp["actual_kg"].tolist()
        if len(vals) < min_history:
            continue
        f = trend_forecast(vals)
        if f is None:
            continue
        row = {}
        key_tuple = key if isinstance(key, tuple) else (key,)
        for c, k in zip(group_cols, key_tuple):
            row[c] = k
        row["forecast_kg"] = round(f, 1)
        row["n_periods_history"] = len(vals)
        rows.append(row)
    return pd.DataFrame(rows)


def compute_shares(sales_df, group_cols, recent_days=120):
    """Historical share of kg for any grouping, from the last `recent_days` (falls back to
    full history if too little recent data). Used to split a single trustworthy total down
    into any breakdown -- guaranteed to sum back to that total exactly, unlike forecasting
    each breakdown independently."""
    d = sales_df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    d = d.dropna(subset=["record_date"])
    if d["record_date"].notna().any():
        cutoff = d["record_date"].max() - pd.Timedelta(days=recent_days)
        recent = d[d["record_date"] >= cutoff]
        d = recent if len(recent) >= 20 else d
    total_kg = d["kg"].sum()
    g = d.groupby(group_cols, as_index=False)["kg"].sum()
    g["share"] = g["kg"] / total_kg if total_kg > 0 else 0
    return g[list(group_cols) + ["share"]]


@st.cache_data(hash_funcs={pd.DataFrame: _cheap_data_fingerprint})
def compute_trending_shares(sales_df, group_cols, freq="W", damping=0.6):
    """Like compute_shares, but projects each segment's share FORWARD based on its own
    recent trend, instead of just averaging history. Real problem this solves: a flat
    average always lags a segment that's genuinely, steadily growing or shrinking its share
    (e.g. a channel taking a growing piece of the business) -- by the time the average
    catches up, it's already behind where the trend is actually heading. Uses the same
    damped growth-rate approach as the rest of the app's forecasting (recent vs prior
    period, growth clamped to +-50%), applied to each segment's share of the total rather
    than to raw volume. Renormalized afterward so every segment's projected share still
    sums to exactly 1.0, and the eventual kg values sum back to the trusted total exactly,
    same guarantee as compute_shares."""
    d = sales_df.copy()
    d["record_date"] = pd.to_datetime(d["record_date"], errors="coerce")
    d = d.dropna(subset=["record_date"])
    if d.empty:
        return pd.DataFrame(columns=list(group_cols) + ["share"])

    if freq == "W":
        d["period"] = (d["record_date"] - pd.to_timedelta(d["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
    else:
        d["period"] = d["record_date"].dt.to_period("M").astype(str)

    period_totals = d.groupby("period")["kg"].sum().rename("period_total")
    seg_period = d.groupby(list(group_cols) + ["period"])["kg"].sum().rename("seg_kg").reset_index()
    seg_period = seg_period.merge(period_totals, on="period")
    seg_period["seg_share"] = seg_period["seg_kg"] / seg_period["period_total"].replace(0, np.nan)

    projected_rows = []
    for key, sub in seg_period.groupby(list(group_cols)):
        sub = sub.sort_values("period")
        share_history = sub["seg_share"].dropna().tolist()
        if len(share_history) < 2:
            projected_share = share_history[-1] if share_history else 0.0
        else:
            n = len(share_history)
            recent = np.median(share_history[-min(4, n):])
            if n >= 8:
                prior = np.median(share_history[-8:-4])
            elif n > min(4, n):
                prior = np.median(share_history[:n - min(4, n)])
            else:
                prior = recent
            growth = 0.0
            if prior > 0:
                growth = max(min((recent - prior) / prior, 1.0), -0.5)
            projected_share = max(recent * (1 + damping * growth), 0.0)
        row = dict(zip(group_cols, key if isinstance(key, tuple) else (key,)))
        row["share"] = projected_share
        projected_rows.append(row)

    result = pd.DataFrame(projected_rows)
    total = result["share"].sum()
    if total > 0:
        result["share"] = result["share"] / total  # renormalize so shares sum to exactly 1.0
    return result


@st.cache_data(show_spinner="Projecting forward...")
def project_forward_with_range(actual_series, error_sigma, n_periods=8):
    """Projects multiple periods ahead. With 2+ years of history (104+ points), tries a real
    seasonal SARIMA(1,1,1)x(1,1,1,52) so the projected range can reflect a genuine yearly
    cycle, not just recent trend. With 8-104 points, non-seasonal ARIMA. Either way, this is
    a single fit producing the whole path at once with real statistical confidence intervals,
    not recursive re-feeding (verified: recursive re-feeding through ARIMA produced a forecast
    that more than doubled over 8 weeks with no plateau; the native multi-step call plateaus
    correctly). Falls back to the damped recursive median method, with an empirical-error-based
    range, when there's too little data for ARIMA. Cached -- ARIMA fits are slow, especially
    the seasonal tier (~12s measured), so this should only run once per actual data change."""
    vals = np.asarray(actual_series, dtype=float)
    n = len(vals)
    if n >= 8:
        seasonal = (1, 1, 1, 52) if n >= 104 else (0, 0, 0, 0)
        try:
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                from statsmodels.tsa.statespace.sarimax import SARIMAX
                model = SARIMAX(vals, order=(1, 1, 1), seasonal_order=seasonal,
                                 enforce_stationarity=False, enforce_invertibility=False)
                fit = model.fit(disp=False)
                result = fit.get_forecast(steps=n_periods)
                mean = result.predicted_mean
                ci = result.conf_int(alpha=0.20)  # ~80% interval, roughly matching a P10-P90 framing
                path = []
                for h in range(n_periods):
                    f = max(float(mean[h]), 0)
                    low = max(float(ci[h, 0]), 0)
                    high = max(float(ci[h, 1]), f)
                    path.append({"step": h + 1, "forecast_kg": f, "low": low, "high": high})
                return pd.DataFrame(path)
        except Exception:
            pass

    hist_vals = list(actual_series)[-8:]
    path = []
    for h in range(1, n_periods + 1):
        effective_damping = 0.6 * (0.7 ** (h - 1))
        f = _median_trend_forecast(hist_vals[-8:], damping=effective_damping)
        if f is None:
            break
        band = error_sigma * np.sqrt(h) if pd.notna(error_sigma) else 0.15 * np.sqrt(h)
        path.append({"step": h, "forecast_kg": f, "low": max(f * (1 - band), 0), "high": f * (1 + band)})
        hist_vals.append(f)
    return pd.DataFrame(path)


# ===================================================================
# GLOBAL STATE
# ===================================================================
st.title("Sales to operations demand planning")
cycle = st.text_input("Planning cycle label", value=current_cycle_label(),
                       help="e.g. 2026-08 -- used for pipeline events and ops capacity/sign-off.")

sales_df = load_sales_records()
has_data = not sales_df.empty

if has_data:
    price_df = compute_price_per_kg(sales_df)
    size_mix_df = compute_size_mix(sales_df)
    customer_mix_df = compute_customer_mix(sales_df)
    weekly_actual = compute_weekly_actuals(sales_df)
    generate_missing_forecasts(weekly_actual)
    backtest_df = backtest_accuracy(weekly_actual)
else:
    price_df = size_mix_df = customer_mix_df = weekly_actual = backtest_df = pd.DataFrame()

# --- current live forecast: latest frozen target week per channel/product ---
live_forecast = pd.read_sql("""
    SELECT af.* FROM auto_forecasts af
    INNER JOIN (SELECT channel, product, MAX(target_week) AS mx FROM auto_forecasts GROUP BY channel, product) t
    ON af.channel=t.channel AND af.product=t.product AND af.target_week=t.mx
""", conn)
if not live_forecast.empty:
    # auto_forecasts is append-only -- if a (channel, product, target_week) combo was ever
    # forecasted more than once (a real bug found and fixed elsewhere: two rows summed to an
    # inflated total), this JOIN would return both rows. Dedupe to the latest by id so every
    # number built from live_forecast (KPIs, the Forecast tab, everything) is never at risk
    # of silently double-counting a duplicate.
    live_forecast = live_forecast.sort_values("id").drop_duplicates(
        subset=["channel", "product", "target_week"], keep="last")

# --- pipeline events, layered on top of the live forecast ---
all_events = pd.read_sql("SELECT * FROM pipeline_events ORDER BY id DESC", conn)
if not all_events.empty:
    applicable = all_events[
        (all_events["starting_cycle"] <= cycle) &
        ((all_events["ongoing"] == 1) | (all_events["starting_cycle"] == cycle))
    ]
    pipeline_by_cp = applicable.groupby(["channel", "product"], as_index=False)["expected_kg_per_month"].sum() \
        .rename(columns={"expected_kg_per_month": "pipeline_kg"})
else:
    applicable = pd.DataFrame()
    pipeline_by_cp = pd.DataFrame(columns=["channel", "product", "pipeline_kg"])

if not live_forecast.empty or not pipeline_by_cp.empty:
    forecast_by_cp = live_forecast[["channel", "product", "forecast_kg", "target_week"]].copy() if not live_forecast.empty \
        else pd.DataFrame(columns=["channel", "product", "forecast_kg", "target_week"])
    forecast_by_cp = forecast_by_cp.merge(pipeline_by_cp, on=["channel", "product"], how="outer")
    forecast_by_cp["forecast_kg"] = forecast_by_cp["forecast_kg"].fillna(0)
    forecast_by_cp["pipeline_kg"] = forecast_by_cp["pipeline_kg"].fillna(0)
    forecast_by_cp["forecast_kg"] = forecast_by_cp["forecast_kg"] + forecast_by_cp["pipeline_kg"]
else:
    forecast_by_cp = pd.DataFrame()

# --- manual overrides: human judgment REPLACES the auto+pipeline number, not adds to it ---
manual_overrides_df = pd.read_sql("SELECT * FROM manual_overrides WHERE active = 1 ORDER BY id DESC", conn)
active_overrides = pd.DataFrame()
if not manual_overrides_df.empty:
    latest_overrides = manual_overrides_df.sort_values("id").groupby(["channel", "product"], as_index=False).last()

    # a "One-time" override only applies while the current live forecast's target_week still
    # matches what it was when the override was set -- once new data moves the target week
    # forward, it's expired, and we auto-deactivate it so it doesn't clutter the active list
    still_valid_rows = []
    for _, row in latest_overrides.iterrows():
        if row["period_type"] == "One-time" and not live_forecast.empty:
            current_target = live_forecast[(live_forecast["channel"] == row["channel"]) &
                                            (live_forecast["product"] == row["product"])]
            if current_target.empty or current_target.iloc[0]["target_week"] != row["target_week"]:
                conn.execute("UPDATE manual_overrides SET active = 0 WHERE id = ?", (int(row["id"]),))
                conn.commit()
                continue
        still_valid_rows.append(row)
    active_overrides = pd.DataFrame(still_valid_rows) if still_valid_rows else pd.DataFrame()

    if not active_overrides.empty:
        if not forecast_by_cp.empty:
            forecast_by_cp = forecast_by_cp.merge(
                active_overrides[["channel", "product", "override_kg"]], on=["channel", "product"], how="outer")
            forecast_by_cp["forecast_kg"] = forecast_by_cp["forecast_kg"].fillna(0)
            forecast_by_cp["pipeline_kg"] = forecast_by_cp["pipeline_kg"].fillna(0)
            forecast_by_cp["forecast_kg"] = np.where(forecast_by_cp["override_kg"].notna(),
                                                      forecast_by_cp["override_kg"], forecast_by_cp["forecast_kg"])
            forecast_by_cp = forecast_by_cp.drop(columns=["override_kg"])
        else:
            forecast_by_cp = active_overrides.rename(columns={"override_kg": "forecast_kg"})
            forecast_by_cp["pipeline_kg"] = 0
            forecast_by_cp["target_week"] = None

# size-level breakdown for ops/bag counts, driven by the auto forecast
if not forecast_by_cp.empty and not size_mix_df.empty:
    translated = forecast_by_cp.merge(size_mix_df[["channel", "product", "size_label", "size_mix_pct"]],
                                       on=["channel", "product"], how="left")
    translated["forecast_kg"] = (translated["forecast_kg"] * translated["size_mix_pct"].fillna(100) / 100).round(1)
    translated = translated[["channel", "product", "size_label", "forecast_kg"]]
else:
    translated = pd.DataFrame()

# implied $ CAD value of the forecast, using the real computed price per kg
if not translated.empty and not price_df.empty:
    dollar_view = translated.merge(price_df[["channel", "product", "size_label", "price_per_kg"]],
                                    on=["channel", "product", "size_label"], how="left")
    dollar_view["forecast_cad"] = (dollar_view["forecast_kg"] * dollar_view["price_per_kg"]).round(2)
    dollar_by_cp = dollar_view.groupby(["channel", "product"], as_index=False).agg(
        forecast_kg=("forecast_kg", "sum"), forecast_cad=("forecast_cad", "sum"))
else:
    dollar_by_cp = pd.DataFrame()


# ===================================================================
# TABS
# ===================================================================
tab_dash, tab_data, tab_rates, tab_forecast, tab_salesplan, tab_pipeline, tab_ops, tab_signoff, tab_ai, tab_history = st.tabs(
    ["Dashboard", "1. Upload sales data", "2. Computed rates", "3. Forecast (auto)",
     "4. Sales plan (S&OP)", "5. Pipeline / known events", "6. Ops capacity check", "7. Sign-off", "8. Ask AI", "9. History"]
)

# --- DASHBOARD (landing page) ---
with tab_dash:
    if not has_data:
        st.info("No sales data uploaded yet. Go to **1. Upload sales data** to get started.")
    elif backtest_df.empty:
        st.warning("Not enough history yet to forecast — need at least a few weeks of data per channel/product.")
    else:
        dim_map = {"Channel": "channel", "Item": "product", "Customer": "customer"}

        # ===============================================================
        # OVERVIEW — always whole-company, never filtered. KPIs + one chart.
        # ===============================================================
        st.markdown("### Overview")

        # Top-down basis: forecast Staple and Single directly (each on its own aggregated
        # series, more statistically reliable than summing many small per-item forecasts --
        # this is standard practice in real demand planning), then this sum becomes the one
        # official total. Falls back to the old bottom-up sum if product types aren't set up.
        type_level_forecasts = compute_type_level_forecast(sales_df, freq="W")

        # attribute each pipeline event's kg impact to ITS OWN product type, not one lump
        # sum floating at the top level -- real gap found: a Staple-item contract wouldn't
        # show up in the Staple panel's own number, only in the overall KPI, meaning Staple
        # + Single would silently stop summing to the KPI total the moment any pipeline
        # event existed. Joining through each product's known classification fixes this the
        # same way everything else this session got reconciled.
        pipeline_by_type = {}
        if not pipeline_by_cp.empty and "product_type" in sales_df.columns:
            product_type_lookup = sales_df[["product", "product_type"]].drop_duplicates()
            pipeline_typed = pipeline_by_cp.merge(product_type_lookup, on="product", how="left")
            pipeline_typed["product_type"] = pipeline_typed["product_type"].fillna("(not tracked)")
            pipeline_by_type = pipeline_typed.groupby("product_type")["pipeline_kg"].sum().to_dict()
        pipeline_total_next_week = pipeline_by_cp["pipeline_kg"].sum() if not pipeline_by_cp.empty else 0

        # each type's forecast now includes ITS OWN attributed pipeline events, so the Staple
        # and Single panels show numbers that are already consistent with the KPI total below
        type_level_forecasts_with_pipeline = {
            pt: val + pipeline_by_type.get(pt, 0) for pt, val in type_level_forecasts.items()
        }
        unattributed_pipeline = pipeline_by_type.get("(not tracked)", 0)  # events on products with no known type

        if type_level_forecasts:
            next_week_kg_all = sum(type_level_forecasts_with_pipeline.values()) + unattributed_pipeline
        else:
            next_week_kg_all = forecast_by_cp["forecast_kg"].sum() if not forecast_by_cp.empty else 0
        next_week_cad_all = dollar_by_cp["forecast_cad"].sum() if not dollar_by_cp.empty else 0

        # the forecast period, in plain dates -- e.g. "Aug 17 - Aug 23, 2026". Derived from
        # LIVE weekly_actual (the same source next_week_kg_all uses), not from the old
        # forecast_by_cp/auto_forecasts lookup -- those are two different things and can
        # genuinely disagree: auto_forecasts is a frozen historical record that's never
        # cleaned up, so deleting a batch of sales data doesn't retroactively remove a
        # forecast that was generated assuming that data existed. Using the live source for
        # both the label and the number keeps them from ever showing an inconsistent pair.
        forecast_period_label = "n/a"
        if not weekly_actual.empty:
            latest_known_week = sorted(weekly_actual["week_start"].unique())[-1]
            tw_start = pd.Timestamp(latest_known_week) + pd.Timedelta(days=7)
            tw_end = tw_start + pd.Timedelta(days=6)
            forecast_period_label = f"{tw_start.strftime('%b %d')} – {tw_end.strftime('%b %d, %Y')}"

        d_kpi = sales_df.copy()
        d_kpi["record_date"] = pd.to_datetime(d_kpi["record_date"], errors="coerce")
        d_kpi = d_kpi.dropna(subset=["record_date"])
        latest_actual_kg = None
        latest_actual_week = None
        if not d_kpi.empty:
            d_kpi["week_start"] = (d_kpi["record_date"] - pd.to_timedelta(d_kpi["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
            wk_kpi = d_kpi.groupby("week_start")["kg"].sum().sort_index()
            if len(wk_kpi):
                latest_actual_kg = wk_kpi.iloc[-1]
                latest_actual_week = wk_kpi.index[-1]

        # last week's forecast accuracy, AND whether it's improving vs the week before --
        # walk-forward recomputed using the CURRENT method (same fix applied to the Staple/
        # Single panel earlier), not a lookup into old frozen auto_forecasts rows, which
        # could show a stale number from a since-replaced forecasting method.
        last_week_accuracy_label, last_week_accuracy_delta = "n/a", None
        if not d_kpi.empty:
            company_weekly = d_kpi.groupby("week_start")["kg"].sum().sort_index()
            weeks_list = company_weekly.index.tolist()

            def _week_accuracy(idx):
                if idx < 2:
                    return None
                history_before = company_weekly.iloc[:idx].tolist()
                f = trend_forecast(history_before)
                actual = company_weekly.iloc[idx]
                if f is None or f <= 0:
                    return None
                pct_off = (actual - f) / f * 100
                return max(0, 100 - abs(pct_off))

            this_week_accuracy = _week_accuracy(len(weeks_list) - 1) if len(weeks_list) >= 1 else None
            prior_week_accuracy = _week_accuracy(len(weeks_list) - 2) if len(weeks_list) >= 2 else None

            if this_week_accuracy is not None:
                last_week_accuracy_label = f"{this_week_accuracy:.0f}% accurate"
                if prior_week_accuracy is not None:
                    trend = this_week_accuracy - prior_week_accuracy
                    last_week_accuracy_delta = f"{trend:+.0f} pts vs the week before"
                else:
                    last_week_accuracy_delta = "no prior week to compare yet"

        cap_row = pd.read_sql("SELECT * FROM ops_capacity WHERE cycle_label = ? ORDER BY id DESC LIMIT 1",
                               conn, params=(cycle,))
        cap_gap = None
        if not cap_row.empty and not forecast_by_cp.empty:
            cap_amt = cap_row.iloc[0]["monthly_capacity_kg"]
            monthly_planned = next_week_kg_all * 4.345
            cap_gap = cap_amt - monthly_planned

        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric(f"Forecast: {forecast_period_label}", f"{next_week_kg_all:,.0f} kg",
                  help="The 7-day period this forecast covers, shown in the label above.")
        k2.metric("Forecast value", f"${next_week_cad_all:,.0f}")
        k3.metric(f"Actual — week of {latest_actual_week}" if latest_actual_week else "Last week actual",
                  f"{latest_actual_kg:,.0f} kg" if latest_actual_kg is not None else "n/a")
        k4.metric("Last week's forecast", last_week_accuracy_label, delta=last_week_accuracy_delta,
                  delta_color="off", help="How close last week's forecast came to what actually happened, "
                                          "and whether that's better or worse than the week before it — "
                                          "so you can tell if accuracy is trending up or down over time.")
        if cap_gap is not None:
            k5.metric("Capacity", "Shortfall" if cap_gap < 0 else "Covered",
                      delta=f"{cap_gap:,.0f} kg/mo", delta_color="normal" if cap_gap >= 0 else "inverse")
        else:
            k5.metric("Capacity", "Not set", help="Set it in tab 5 to see a shortfall check here.")

        st.caption("Solid blue line is **actual** historical sales, whole company. Dashed line (Week view "
                   "only) is what the auto-forecast would have predicted for each of the last 12 weeks, "
                   "checked against what actually happened — a real accuracy check, not a future prediction. "
                   "Tick the box below to also see the forecast projected forward, with a range built from "
                   "this model's own real historical accuracy.")
        trend_freq = st.radio("Show by", ["Week", "Month"], horizontal=True, key="trend_freq")
        d_trend = sales_df.copy()
        d_trend["record_date"] = pd.to_datetime(d_trend["record_date"], errors="coerce")
        d_trend = d_trend.dropna(subset=["record_date"])
        if trend_freq == "Week":
            d_trend["period"] = (d_trend["record_date"] - pd.to_timedelta(d_trend["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
        else:
            d_trend["period"] = d_trend["record_date"].dt.to_period("M").astype(str)
        trend_agg = d_trend.groupby("period", as_index=False)["kg"].sum().sort_values("period")

        total_bt = backtest_df.groupby("week_start", as_index=False).agg(
            forecast_kg=("forecast_kg", "sum"), actual_kg=("actual_kg", "sum"))
        total_bt["variance_pct"] = (total_bt["actual_kg"] - total_bt["forecast_kg"]) / total_bt["forecast_kg"].replace(0, np.nan)
        error_sigma = total_bt["variance_pct"].std()

        n_periods_fwd = 8 if trend_freq == "Week" else 6
        show_projection = st.checkbox("Also show the forecast projection (runs a real seasonal fit, "
                                       "~15-20s first time, cached after)", key="show_trend_projection")
        if show_projection:
            with st.spinner("Projecting the trend forward..."):
                projection = project_forward_with_range(trend_agg["kg"].tolist(), error_sigma, n_periods=n_periods_fwd)
        else:
            projection = pd.DataFrame()

        fig_trend = go.Figure()
        # display only the last ~6 months on the chart (recent, relevant view) -- the full
        # history is still used for the underlying error_sigma/projection math above, this
        # is purely a display trim, not a computation change
        n_display = 26 if trend_freq == "Week" else 6
        trend_agg_display = trend_agg.tail(n_display)
        fig_trend.add_trace(go.Scatter(x=trend_agg_display["period"], y=trend_agg_display["kg"], mode="lines", name="Actual",
                                        line=dict(color="rgb(31,119,180)", width=2)))
        if trend_freq == "Week" and not total_bt.empty:
            bt_recent = total_bt.sort_values("week_start").tail(12)
            fig_trend.add_trace(go.Scatter(x=bt_recent["week_start"], y=bt_recent["forecast_kg"], mode="lines",
                                            name="Auto forecast (backtested)",
                                            line=dict(color="rgb(139,90,60)", width=2, dash="dash")))
        if not projection.empty:
            last_date = pd.Timestamp(trend_agg["period"].iloc[-1])
            if trend_freq == "Week":
                future_dates = [(last_date + pd.Timedelta(weeks=int(s))).date().isoformat() for s in projection["step"]]
            else:
                future_dates = [(last_date + pd.DateOffset(months=int(s))).date().isoformat() for s in projection["step"]]
            join_x = [trend_agg["period"].iloc[-1]] + future_dates
            join_y_high = [trend_agg["kg"].iloc[-1]] + projection["high"].tolist()
            join_y_low = [trend_agg["kg"].iloc[-1]] + projection["low"].tolist()
            join_y_mid = [trend_agg["kg"].iloc[-1]] + projection["forecast_kg"].tolist()
            fig_trend.add_trace(go.Scatter(x=join_x, y=join_y_high, mode="lines", line=dict(width=0),
                                            showlegend=False, hoverinfo="skip"))
            fig_trend.add_trace(go.Scatter(x=join_x, y=join_y_low, mode="lines", line=dict(width=0),
                                            fill="tonexty", fillcolor="rgba(120,120,120,0.25)", name="Forecast range",
                                            hoverinfo="skip"))
            fig_trend.add_trace(go.Scatter(x=join_x, y=join_y_mid, mode="lines", name="Forecast",
                                            line=dict(color="rgb(60,60,60)", width=2)))
        fig_trend.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10), plot_bgcolor="white",
                                 xaxis_title=trend_freq, yaxis_title="Total kg (all channels/items)",
                                 hovermode="x unified", xaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.06)"),
                                 yaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.06)"))
        st.plotly_chart(fig_trend, use_container_width=True)

        st.divider()

        # ===============================================================
        # SIMPLE FORECAST — by Single vs Staple. Deliberately minimal --
        # built for a quick manager glance, not detailed analysis.
        # ===============================================================
        product_types_available = sorted(
            sales_df[sales_df["product_type"] != "(not tracked)"]["product_type"].unique().tolist()
        ) if "product_type" in sales_df.columns and (sales_df["product_type"] != "(not tracked)").any() else []

        if product_types_available:
            st.markdown("## Staple vs Single")
            st.caption("Each forecasted independently from its own history — not derived by "
                       "splitting the total after the fact. Their sum is what makes up the "
                       "Overview total above, so they always agree with it by construction.")
            pt_horizon = st.radio("Show by", ["Week", "Month"], horizontal=True, key="pt_horizon")
            n_periods_shown = 8 if pt_horizon == "Week" else 6
            n_history_shown = 6 if pt_horizon == "Week" else 6
            freq = "W" if pt_horizon == "Week" else "M"

            type_level_forecasts_h = compute_type_level_forecast(sales_df, freq=freq)
            if freq == "W" and pipeline_by_type:
                # same pipeline attribution as the KPI above -- keeps this panel's own numbers
                # consistent with the Overview total, including any logged pipeline events
                type_level_forecasts_h = {pt: val + pipeline_by_type.get(pt, 0) for pt, val in type_level_forecasts_h.items()}

            pt_cols = st.columns(len(product_types_available))
            for idx, pt in enumerate(product_types_available):
                with pt_cols[idx]:
                    st.markdown(f"### {pt}")
                    pt_df = sales_df[sales_df["product_type"] == pt]

                    agg_pt = aggregate_periods(pt_df, ["product_type"], freq)
                    agg_pt = agg_pt.sort_values("period")

                    if len(agg_pt) < 2:
                        st.info("Not enough history yet for this category.")
                        continue

                    with st.spinner(f"Projecting {pt} forward (first time can take a moment, cached after)..."):
                        projection_pt = project_forward_with_range(agg_pt["actual_kg"].tolist(), None, n_periods=n_periods_shown)

                    # anchor period 1 to the SAME single-step forecast that feeds the Overview
                    # KPI total -- guarantees this table's first period matches the top of the
                    # dashboard exactly. Uses an ADDITIVE offset, not a proportional rescale of
                    # the whole curve -- real bug found and fixed: a proportional rescale means
                    # a one-time pipeline event (meant to expire after the current month) was
                    # silently leaking its effect, proportionally, all the way out to week 8.
                    # An additive offset to period 1 only keeps a one-time bump contained to the
                    # period it actually applies to, leaving the statistical shape of later
                    # periods untouched.
                    direct_forecast = type_level_forecasts_h.get(pt)
                    if not projection_pt.empty and direct_forecast is not None:
                        offset = direct_forecast - projection_pt["forecast_kg"].iloc[0]
                        projection_pt.loc[projection_pt.index[0], "forecast_kg"] = max(direct_forecast, 0)
                        projection_pt.loc[projection_pt.index[0], "low"] = max(projection_pt["low"].iloc[0] + offset, 0)
                        projection_pt.loc[projection_pt.index[0], "high"] = max(projection_pt["high"].iloc[0] + offset, 0)

                    total_company_this_period = sum(type_level_forecasts_h.values()) if type_level_forecasts_h else None
                    if total_company_this_period:
                        st.caption(f"{(direct_forecast or 0)/total_company_this_period*100:.0f}% of next period's total")

                    # simple table: recent actual -> forecast, one continuous timeline
                    recent_actual = agg_pt.tail(n_history_shown)[["period", "actual_kg"]].rename(
                        columns={"period": "Period", "actual_kg": "Actual (kg)"})

                    if pt_horizon == "Week":
                        # walk-forward recompute using the CURRENT top-down method (same one
                        # used for the forward projection above), rather than looking up old
                        # frozen auto_forecasts rows. Real issue found: those rows were frozen
                        # by the OLD per-item bottom-up method, before this app was rebuilt to
                        # forecast Staple/Single directly -- so a week could show a stored value
                        # from a since-replaced method once it became historical, inconsistent
                        # with the live projection shown for it a moment earlier as a future
                        # week. Recomputing fresh with today's method for each week (using only
                        # data that existed before that week -- no peeking) keeps this genuinely
                        # consistent and honest, matching the walk-forward validation approach.
                        agg_pt_sorted = agg_pt.sort_values("period").reset_index(drop=True)
                        walk_forward_rows = []
                        for i in range(len(agg_pt_sorted) - n_history_shown, len(agg_pt_sorted)):
                            if i < 2:
                                continue
                            history_before = agg_pt_sorted["actual_kg"].iloc[:i].tolist()
                            f = trend_forecast(history_before)  # fast method -- several of these run per render
                            walk_forward_rows.append({"Period": agg_pt_sorted["period"].iloc[i], "Forecast (kg)": f})
                        stored_by_week = pd.DataFrame(walk_forward_rows) if walk_forward_rows \
                            else pd.DataFrame(columns=["Period", "Forecast (kg)"])
                        recent_actual = recent_actual.merge(stored_by_week, on="Period", how="left")
                    else:
                        recent_actual["Forecast (kg)"] = None

                    fwd_table = projection_pt.copy()
                    last_date = pd.Timestamp(agg_pt["period"].iloc[-1])
                    if pt_horizon == "Week":
                        fwd_table["Period"] = [(last_date + pd.Timedelta(weeks=int(s))).date().isoformat() for s in fwd_table["step"]]
                    else:
                        fwd_table["Period"] = [(last_date + pd.DateOffset(months=int(s))).date().isoformat() for s in fwd_table["step"]]
                    fwd_table["Actual (kg)"] = None
                    fwd_table = fwd_table.rename(columns={"forecast_kg": "Forecast (kg)"})[["Period", "Actual (kg)", "Forecast (kg)"]]

                    # don't duplicate a period that already has a stored forecast merged in above
                    fwd_table = fwd_table[~fwd_table["Period"].isin(recent_actual["Period"])]

                    simple_table = pd.concat([recent_actual, fwd_table], ignore_index=True)
                    simple_table["Actual (kg)"] = simple_table["Actual (kg)"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                    simple_table["Forecast (kg)"] = simple_table["Forecast (kg)"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                    st.dataframe(simple_table.set_index("Period").T, use_container_width=True)
                    if pt_horizon == "Month":
                        st.caption("Monthly view only tracks live forecasts going forward — historical "
                                   "forecast-vs-actual by month isn't stored yet (weekly view has it).")

                    fig_pt = go.Figure()
                    # focus the x-axis on the last 6 months, same as the Overview chart -- the
                    # full history still feeds the actual forecasting math above, this only
                    # trims what's displayed
                    n_display_pt = 26 if pt_horizon == "Week" else 6
                    agg_pt_display = agg_pt.tail(n_display_pt)
                    fig_pt.add_trace(go.Scatter(x=agg_pt_display["period"], y=agg_pt_display["actual_kg"], mode="lines",
                                                 name="Actual", line=dict(color="rgb(31,119,180)", width=2)))
                    hist_forecast = recent_actual.dropna(subset=["Forecast (kg)"])
                    if not hist_forecast.empty:
                        fig_pt.add_trace(go.Scatter(x=hist_forecast["Period"], y=hist_forecast["Forecast (kg)"],
                                                     mode="lines+markers", name="Forecast (checked against actual)",
                                                     line=dict(color="rgb(139,90,60)", width=2, dash="dash")))
                    if not projection_pt.empty:
                        join_x_pt = [agg_pt["period"].iloc[-1]] + fwd_table["Period"].tolist()
                        join_y_pt = [agg_pt["actual_kg"].iloc[-1]] + projection_pt["forecast_kg"].tolist()
                        fig_pt.add_trace(go.Scatter(x=join_x_pt, y=join_y_pt, mode="lines", name="Forecast (ahead)",
                                                     line=dict(color="rgb(60,60,60)", width=2)))
                    fig_pt.update_layout(height=240, margin=dict(l=10, r=10, t=10, b=10), plot_bgcolor="white",
                                          showlegend=True, xaxis=dict(showgrid=False), yaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.06)"))
                    st.plotly_chart(fig_pt, use_container_width=True)

            st.divider()

        # ===============================================================
        # STAPLE — CHANNEL & BAG SIZE, 3 MONTHS AHEAD (for Operations' bag ordering)
        # ===============================================================
        if "Staple" in product_types_available and "size_label" in sales_df.columns:
            st.markdown("## Staple — by channel & bag size, 3 months ahead")
            st.caption(
                "For Operations' bag ordering, which needs roughly a 3-month lead time. Specialty Retail "
                "is forecasted directly (it's grown to ~46%+ of Staple and is still actively shifting, so "
                "a proportional split would lag its real trajectory) — every other channel splits the "
                "remainder by trending share, and bag size within each channel the same way. If a size "
                "looks off here, that's the signal to order early rather than find out in 3 months."
            )
            if st.button("Compute breakdown", key="compute_staple_breakdown"):
                st.session_state["show_staple_breakdown"] = True

            if st.session_state.get("show_staple_breakdown"):
                breakdown_df = compute_staple_channel_breakdown(sales_df, n_periods=3, freq="M")
                if breakdown_df.empty:
                    st.info("Not enough Staple history yet for this breakdown.")
                elif breakdown_df["forecast_bags"].isna().all():
                    st.warning("Showing kg only — no quantity/bag-count data was captured on upload yet, "
                               "so a real bag count can't be computed. Re-upload with a Quantity column "
                               "mapped (tab 1) to get actual bag counts here.")
                    pivot = breakdown_df.pivot_table(index=["channel", "size_label"], columns="period",
                                                       values="forecast_kg", aggfunc="sum").round(0)
                    st.dataframe(pivot, use_container_width=True)
                else:
                    show_unit = st.radio("Show as", ["Bags (for ordering)", "Kg"], horizontal=True, key="staple_bag_unit")
                    value_col = "forecast_bags" if show_unit.startswith("Bags") else "forecast_kg"
                    pivot = breakdown_df.pivot_table(index=["channel", "size_label"], columns="period",
                                                       values=value_col, aggfunc="sum").round(0)
                    st.dataframe(pivot, use_container_width=True)
                    with st.expander("Kg-per-bag rates used (learned from your actual data)"):
                        st.dataframe(compute_kg_per_bag(sales_df), use_container_width=True, hide_index=True)
                    st.caption("Rows sum to each channel's reconciled total; every channel's total sums to "
                               "the overall Staple forecast for that month.")
            st.divider()

        st.markdown("### Filter / break down by")
        st.caption("'Not included' aggregates across every value of that dimension (e.g. item totals combined "
                   "across all channels). 'All' breaks that dimension down into every value (rows in tables, "
                   "bars in charts). Pick one specific value to narrow everything below to just that segment.")
        fc1, fc2, fc3 = st.columns(3)
        channel_options = ["(not included)", "All"] + sorted(sales_df["channel"].unique().tolist())
        item_options = ["(not included)", "All"] + sorted(sales_df["product"].unique().tolist())
        customer_available = "customer" in sales_df.columns and not (sales_df["customer"] == "(not tracked)").all()
        customer_options = ["(not included)", "All"] + sorted(sales_df[sales_df["customer"] != "(not tracked)"]["customer"].unique().tolist()) \
            if customer_available else ["(not included)"]

        sel_channel = fc1.selectbox("Channel", channel_options, key="filt_channel", index=1)
        sel_item = fc2.selectbox("Item", item_options, key="filt_item", index=1)
        sel_customer = fc3.selectbox("Customer", customer_options, key="filt_customer", disabled=not customer_available,
                                      help="'Not included' aggregates across all customers, without breaking down "
                                           "or filtering — 'All' adds it as a breakdown dimension (sparse accounts "
                                           "get filtered out automatically), or pick one specific customer.")
        if not customer_available:
            fc3.caption("Not available in this data source.")

        group_cols = []
        filter_values = {}
        if sel_channel == "All":
            group_cols.append("channel")
        elif sel_channel not in ("(not included)",):
            filter_values["channel"] = sel_channel
        if sel_item == "All":
            group_cols.append("product")
        elif sel_item not in ("(not included)",):
            filter_values["product"] = sel_item
        if sel_customer == "All":
            group_cols.append("customer")
        elif sel_customer not in ("(not included)",):
            filter_values["customer"] = sel_customer

        filtered_df = sales_df.copy()
        for col, val in filter_values.items():
            filtered_df = filtered_df[filtered_df[col] == val]

        specific_share = 1.0
        if filter_values:
            if filtered_df.empty:
                st.warning("No records match this combination — try a different filter.")
                specific_share = 0.0
            else:
                share_base = sales_df[sales_df["customer"] != "(not tracked)"] if "customer" in filter_values else sales_df
                filt_shares = compute_shares(share_base, list(filter_values.keys()))
                match = filt_shares.copy()
                for col, val in filter_values.items():
                    match = match[match[col] == val]
                specific_share = match["share"].sum() if not match.empty else 0.0
                st.caption(f"Filtering to: {', '.join(f'{k}={v}' for k, v in filter_values.items())} "
                           f"— {specific_share*100:.1f}% of total company volume.")

        next_week_kg = next_week_kg_all * specific_share
        next_week_cad = next_week_cad_all * specific_share

        st.divider()

        # ===============================================================
        # DETAILED — everything below responds to the filter above
        # ===============================================================
        st.markdown("## Detailed breakdown")

        # --- forecast ahead ---
        st.markdown("### Forecast ahead")
        st.caption("The total is always computed once, from the full channel × item history — every "
                   "breakdown below is that same total split by real historical share, so they always "
                   "add up to the same number no matter how you slice it.")
        horizon = st.radio("Horizon", ["Next week", "Next month", "Next year"], horizontal=True, key="fwd_horizon")

        if horizon == "Next week":
            # exactly the same authoritative number as the Overview KPI above -- includes pipeline
            # events and manual overrides, not just the raw statistical forecast
            canonical_total = next_week_kg_all
        else:
            agg_canonical = aggregate_periods(sales_df, ["channel", "product"], "M")
            canonical_fwd = forecast_next_period(agg_canonical, ["channel", "product"], min_history=2)
            canonical_total = canonical_fwd["forecast_kg"].sum() if not canonical_fwd.empty else 0
            if horizon == "Next year":
                canonical_total = canonical_total * 12
            st.caption("Note: unlike 'Next week', this month/year projection doesn't yet include Pipeline "
                       "events or Manual overrides — it's the raw statistical forecast only.")
        canonical_total = canonical_total * specific_share

        if not group_cols:
            if filter_values:
                st.info("All three are filtered to specific values, so there's nothing left to break down — "
                        "check the KPI cards and Overall Trend above for this exact segment.")
            else:
                st.info("No dimension is set to 'All', so there's nothing to break down into rows — "
                        "set at least one to 'All' to see a breakdown, or check Overview above for the total.")
        elif canonical_total == 0:
            st.info("Not enough history yet to forecast.")
        else:
            needs_customer_fwd = "customer" in group_cols
            if needs_customer_fwd and (not has_data or "customer" not in sales_df.columns
                                        or (sales_df["customer"] == "(not tracked)").all()):
                st.warning("This data source doesn't include customer identity, so a breakdown "
                           "including Customer isn't available.")
                shares = pd.DataFrame()
            else:
                base_df_fwd = filtered_df[filtered_df["customer"] != "(not tracked)"] if needs_customer_fwd else filtered_df
                shares = compute_shares(base_df_fwd, group_cols)

            if shares.empty:
                st.info("Not enough history yet for this breakdown.")
            else:
                shares["forecast_kg"] = (shares["share"] * canonical_total).round(1)
                shares["Segment"] = shares[group_cols].astype(str).agg(" — ".join, axis=1) \
                    if len(group_cols) > 1 else shares[group_cols[0]]
                period_label = {"Next week": "Forecast kg (next week)", "Next month": "Forecast kg (next month)",
                                 "Next year": "Forecast kg (next year, extrapolated)"}[horizon]

                st.metric(f"Total — {horizon.lower()}", f"{canonical_total:,.0f} kg")
                display_shares = shares.sort_values("forecast_kg", ascending=False)
                st.dataframe(
                    display_shares[["Segment", "forecast_kg"]].rename(columns={"forecast_kg": period_label}),
                    use_container_width=True, hide_index=True)
                if horizon == "Next year":
                    st.caption("Next year isn't independently modeled — not enough history yet (multiple "
                               "full years) for a real year-over-year trend. This is the monthly total × "
                               "12, a simple extrapolation. Treat it as a rough planning figure.")

        if not dollar_by_cp.empty:
            with st.expander("Translated forecast — kg and CAD, by channel and item"):
                display_dollar = dollar_by_cp.copy()
                display_dollar["forecast_kg"] = display_dollar["forecast_kg"].round(0)
                display_dollar["forecast_cad"] = display_dollar["forecast_cad"].map(lambda x: f"${x:,.0f}")
                display_dollar = display_dollar.rename(columns={
                    "channel": "Channel", "product": "Item", "forecast_kg": "Forecast (kg)", "forecast_cad": "Forecast (CAD)"})
                st.dataframe(display_dollar, use_container_width=True, hide_index=True)
                c1, c2 = st.columns(2)
                c1.metric("Combined total — kg", f"{next_week_kg:,.0f} kg")
                c2.metric("Combined total — CAD", f"${next_week_cad:,.0f}")
                st.caption("CAD value is the forecast kg × the real weighted price/kg computed from your "
                           "uploaded sales history — not a manually entered number.")

        st.divider()
        view = st.radio("View", ["Weekly report", "Monthly report"], horizontal=True)

        if view == "Weekly report":
            st.markdown("**Accuracy overview — every segment at a glance**")
            st.caption(f"Broken down by: {' × '.join(group_cols) if group_cols else '(none — fully filtered to one specific segment)'}")

            if not group_cols:
                if filter_values:
                    st.info("All three are filtered to specific values, so there's nothing left to break down — "
                            "check the KPI cards and Overall Trend above for this exact segment.")
                else:
                    st.info("No dimension is set to 'All', so there's nothing to break down into rows — "
                            "set at least one to 'All' to see a breakdown, or check Overview above for the total.")
                bt = pd.DataFrame()
                cadence_df = pd.DataFrame()
            elif "customer" in group_cols:
                cadence_df = pd.DataFrame()
                if not has_data or "customer" not in sales_df.columns or (sales_df["customer"] == "(not tracked)").all():
                    st.warning("This data source doesn't include customer identity, so any grouping including "
                               "Customer isn't available.")
                    bt = pd.DataFrame()
                else:
                    base_df = sales_df[sales_df["customer"] != "(not tracked)"]
                    weekly_g = compute_weekly_actuals_by(base_df, group_cols)
                    bt = backtest_accuracy(weekly_g, group_cols=group_cols)
                    if not bt.empty:
                        enough_history = bt.groupby(group_cols, as_index=False)["week_start"].count()
                        keep_df = enough_history[enough_history["week_start"] >= 3][group_cols]
                        bt = bt.merge(keep_df, on=group_cols, how="inner")
                        if bt.empty:
                            st.info("No combination has enough order history yet (need 3+ forecastable weeks) "
                                     "at this granularity — try removing Customer or a dimension.")
                        elif group_cols == ["customer"]:
                            cadence_rows = []
                            latest_data_date = pd.to_datetime(sales_df["record_date"], errors="coerce").max()
                            for cust, grp in weekly_g[weekly_g["customer"].isin(bt["customer"].unique())].groupby("customer"):
                                dates = pd.to_datetime(grp["week_start"]).sort_values()
                                gaps = dates.diff().dt.days.dropna() / 7
                                cadence_rows.append({
                                    "customer": cust,
                                    "avg_reorder_weeks": round(gaps.mean(), 1) if len(gaps) else None,
                                    "weeks_since_last_order": round((latest_data_date - dates.max()).days / 7, 1),
                                })
                            cadence_df = pd.DataFrame(cadence_rows)
            else:
                cadence_df = pd.DataFrame()
                bt = backtest_df.copy()
                if group_cols != ["channel", "product"] and not bt.empty:
                    bt = bt.groupby(group_cols + ["week_start"], as_index=False).agg(
                        forecast_kg=("forecast_kg", "sum"), actual_kg=("actual_kg", "sum"),
                        n_weeks_history=("n_weeks_history", "min"))
                    bt["variance_pct"] = (bt["actual_kg"] - bt["forecast_kg"]) / bt["forecast_kg"].replace(0, np.nan)

            if bt.empty:
                st.info("Nothing to show for this grouping yet.")
            else:
                latest_week = bt["week_start"].max()
                overview_rows = []
                for key, grp in bt.groupby(group_cols):
                    grp = grp.sort_values("week_start")
                    label = key if isinstance(key, str) else " — ".join(key)
                    last_row = grp.iloc[-1]
                    recent_bias = grp["variance_pct"].tail(4).mean()
                    weeks_of_history = grp["n_weeks_history"].iloc[-1] if "n_weeks_history" in grp.columns else None
                    if pd.isna(recent_bias):
                        status = "Not enough data"
                    elif abs(recent_bias) > 0.15:
                        status = "ALERT"
                    elif abs(recent_bias) > 0.08:
                        status = "WATCH"
                    else:
                        status = "OK"
                    confidence = "Low (little history)" if (weeks_of_history is not None and weeks_of_history < 4) else "Normal"
                    row = {
                        "Segment": label, "Latest forecast (kg)": round(last_row["forecast_kg"]),
                        "Latest actual (kg)": round(last_row["actual_kg"]),
                        "Recent 4wk bias": f"{recent_bias*100:+.0f}%" if pd.notna(recent_bias) else "n/a",
                        "Confidence": confidence, "Status": status,
                    }
                    if group_cols == ["customer"] and not cadence_df.empty:
                        cad = cadence_df[cadence_df["customer"].astype(str) == str(label)]
                        if not cad.empty:
                            row["Avg reorder (weeks)"] = cad.iloc[0]["avg_reorder_weeks"]
                            row["Weeks since last order"] = cad.iloc[0]["weeks_since_last_order"]
                    overview_rows.append(row)
                overview_df = pd.DataFrame(overview_rows).sort_values(
                    "Status", key=lambda s: s.map({"ALERT": 0, "WATCH": 1, "OK": 2, "Not enough data": 3}))

                def _flag(row):
                    color = {"ALERT": "background-color: #fbeae6", "WATCH": "background-color: #fdf3e0",
                             "OK": "", "Not enough data": ""}.get(row["Status"], "")
                    return [color] * len(row)
                st.dataframe(overview_df.style.apply(_flag, axis=1), use_container_width=True, hide_index=True)
                st.caption(f"As of week of {latest_week}. ALERT = recent actuals off by 15%+ from forecast, "
                           "WATCH = 8-15%. Confidence 'Low' means under 4 weeks of history fed the forecast — "
                           "treat those numbers as rough, not reliable.")

        else:  # Monthly report
            bt = backtest_df.copy()
            bt["month"] = pd.to_datetime(bt["week_start"]).dt.to_period("M").astype(str)
            bt["abs_variance_pct"] = bt["variance_pct"].abs()
            monthly = bt.groupby(["channel", "product", "month"], as_index=False).agg(
                MAPE=("abs_variance_pct", "mean"), Bias=("variance_pct", "mean"), weeks=("week_start", "count"))
            monthly["MAPE_%"] = (monthly["MAPE"] * 100).round(1)
            monthly["Bias_%"] = (monthly["Bias"] * 100).round(1)
            st.dataframe(monthly[["month", "channel", "product", "MAPE_%", "Bias_%", "weeks"]]
                         .sort_values("month", ascending=False), use_container_width=True)
            st.caption("Positive bias = actuals running ahead of the auto-forecast (under-forecasting). "
                       "Negative = over-forecasting. MAPE = average error size regardless of direction.")

    # --- downloadable snapshot report, for meetings ---
    if has_data and not backtest_df.empty:
        st.divider()
        st.subheader("Download a report")
        st.caption("A self-contained snapshot of the current dashboard — everyone in a meeting can open it, "
                   "no login or app access needed.")

        # shared data prep, used by all three report formats
        gen_time = datetime.now().strftime("%Y-%m-%d %H:%M")
        cap_row2 = pd.read_sql("SELECT * FROM ops_capacity WHERE cycle_label = ? ORDER BY id DESC LIMIT 1",
                                conn, params=(cycle,))
        cap_status_text, cap_shortfall = None, False
        if not cap_row2.empty and not forecast_by_cp.empty:
            cap2 = cap_row2.iloc[0]["monthly_capacity_kg"]
            planned2 = forecast_by_cp["forecast_kg"].sum() * 4.345
            cap_shortfall = cap2 < planned2
            cap_status_text = (f"Capacity check: {planned2:,.0f} kg/month planned vs {cap2:,.0f} kg/month "
                                f"capacity — {'SHORTFALL' if cap_shortfall else 'OK'}")

        report_dollar_df = pd.DataFrame()
        if not dollar_by_cp.empty:
            report_dollar_df = dollar_by_cp.rename(columns={"channel": "Channel", "product": "Item",
                                                              "forecast_kg": "Forecast (kg)", "forecast_cad": "Forecast (CAD)"})

        ov = backtest_df.copy()
        latest_wk = ov["week_start"].max()
        summary_rows = []
        for key, grp in ov.groupby(["channel", "product"]):
            grp = grp.sort_values("week_start")
            bias = grp["variance_pct"].tail(4).mean()
            status = "ALERT" if pd.notna(bias) and abs(bias) > 0.15 else \
                ("WATCH" if pd.notna(bias) and abs(bias) > 0.08 else ("OK" if pd.notna(bias) else "n/a"))
            summary_rows.append({"Channel": key[0], "Item": key[1],
                                  "Latest forecast (kg)": round(grp.iloc[-1]["forecast_kg"]),
                                  "Latest actual (kg)": round(grp.iloc[-1]["actual_kg"]),
                                  "Recent 4wk bias": f"{bias*100:+.0f}%" if pd.notna(bias) else "n/a",
                                  "Status": status})
        report_overview_df = pd.DataFrame(summary_rows).sort_values(
            "Status", key=lambda s: s.map({"ALERT": 0, "WATCH": 1, "OK": 2, "n/a": 3}))

        acc = backtest_df.copy()
        acc["abs_variance_pct"] = acc["variance_pct"].abs()
        acc_monthly = acc.copy()
        acc_monthly["month"] = pd.to_datetime(acc_monthly["week_start"]).dt.to_period("M").astype(str)
        report_accuracy_df = acc_monthly.groupby(["channel", "product"], as_index=False).agg(
            MAPE=("abs_variance_pct", "mean"), Bias=("variance_pct", "mean"), weeks_tracked=("week_start", "count"))
        report_accuracy_df["MAPE_%"] = (report_accuracy_df["MAPE"] * 100).round(1)
        report_accuracy_df["Bias_%"] = (report_accuracy_df["Bias"] * 100).round(1)
        report_accuracy_df = report_accuracy_df.rename(columns={"channel": "Channel", "product": "Item"})[
            ["Channel", "Item", "MAPE_%", "Bias_%", "weeks_tracked"]].sort_values("MAPE_%", ascending=False)

        # trend chart data (actual, all history) for embedding
        d_report_trend = sales_df.copy()
        d_report_trend["record_date"] = pd.to_datetime(d_report_trend["record_date"], errors="coerce")
        d_report_trend = d_report_trend.dropna(subset=["record_date"])
        d_report_trend["period"] = (d_report_trend["record_date"] -
                                     pd.to_timedelta(d_report_trend["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
        report_trend_df = d_report_trend.groupby("period", as_index=False)["kg"].sum().sort_values("period")

        report_bar_df = dollar_by_cp.copy() if not dollar_by_cp.empty else pd.DataFrame()

        def build_report_html():
            cap_html = ""
            if cap_status_text:
                color2 = "#b3432f" if cap_shortfall else "#4a7a5c"
                cap_html = f'<p style="color:{color2};font-weight:600">{cap_status_text}</p>'
            if not report_dollar_df.empty:
                display_df = report_dollar_df.copy()
                display_df["Forecast (CAD)"] = display_df["Forecast (CAD)"].map(lambda x: f"${x:,.0f}")
                dollar_html = display_df.to_html(index=False, border=0)
            else:
                dollar_html = "<p>No forecast available yet.</p>"
            overview_html = report_overview_df.to_html(index=False, border=0)
            accuracy_html = report_accuracy_df.to_html(index=False, border=0) if not report_accuracy_df.empty else "<p>Not enough history yet.</p>"

            trend_fig = go.Figure(go.Scatter(x=report_trend_df["period"], y=report_trend_df["kg"], mode="lines",
                                              fill="tozeroy", line=dict(color="rgb(139,90,60)", width=2)))
            trend_fig.update_layout(height=280, margin=dict(l=10, r=10, t=10, b=10),
                                     xaxis_title="Week", yaxis_title="Total actual kg")
            trend_chart_html = trend_fig.to_html(include_plotlyjs="cdn", full_html=False)

            bar_chart_html = ""
            if not report_bar_df.empty:
                bar_sorted = report_bar_df.sort_values("forecast_kg", ascending=True).tail(15)
                bar_fig = go.Figure(go.Bar(x=bar_sorted["forecast_kg"],
                                            y=bar_sorted["channel"] + " — " + bar_sorted["product"],
                                            orientation="h", marker_color="rgb(139,90,60)"))
                bar_fig.update_layout(height=max(280, 24 * len(bar_sorted)), margin=dict(l=10, r=10, t=10, b=10),
                                       xaxis_title="Forecast kg")
                bar_chart_html = bar_fig.to_html(include_plotlyjs=False, full_html=False)

            return f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<title>Demand Planning Report — {cycle}</title>
<style>
body{{font-family:-apple-system,Arial,sans-serif;max-width:900px;margin:2rem auto;color:#2b2622;padding:0 1.5rem}}
h1{{font-size:22px;margin-bottom:4px}} h2{{font-size:16px;margin-top:2rem}}
table{{width:100%;border-collapse:collapse;font-size:13px;margin-top:0.5rem}}
th,td{{text-align:left;padding:6px 10px;border-bottom:1px solid #e3ddd1}}
th{{background:#f3efe8}} .meta{{color:#6b6258;font-size:13px}}
</style></head><body>
<h1>Demand Planning Report</h1>
<p class="meta">Cycle {cycle} — generated {gen_time} — as of week of {latest_wk}</p>
{cap_html}
<h2>Overall trend — actual sales</h2>
{trend_chart_html}
<h2>Translated forecast — kg and CAD</h2>
{dollar_html}
<h2>Forecast by segment</h2>
{bar_chart_html if bar_chart_html else "<p>No forecast available yet.</p>"}
<h2>Forecast accuracy overview</h2>
{overview_html}
<h2>Forecast accuracy — MAPE and bias</h2>
{accuracy_html}
<p class="meta">ALERT = actuals off by 15%+ from forecast over the last 4 weeks. WATCH = 8-15%.
MAPE = average error size regardless of direction. Bias: positive = under-forecasting, negative = over-forecasting.
Generated automatically from 49th Parallel's demand planning app.</p>
</body></html>"""

        def build_report_pdf():
            def safe(t):
                t = str(t)
                for a, b in [("\u2014", "-"), ("\u2013", "-"), ("\u2018", "'"), ("\u2019", "'"),
                             ("\u201c", '"'), ("\u201d", '"'), ("\u2026", "...")]:
                    t = t.replace(a, b)
                return t.encode("latin-1", "replace").decode("latin-1")

            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 16)
            pdf.cell(0, 10, "Demand Planning Report", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(0, 6, safe(f"Cycle {cycle} -- generated {gen_time} -- as of week of {latest_wk}"),
                     new_x="LMARGIN", new_y="NEXT")
            if cap_status_text:
                pdf.ln(2)
                pdf.set_text_color(179, 67, 47) if cap_shortfall else pdf.set_text_color(74, 122, 92)
                pdf.set_font("Helvetica", "B", 10)
                pdf.multi_cell(0, 6, safe(cap_status_text))
                pdf.set_text_color(0, 0, 0)

            def draw_table(title, df, col_widths):
                pdf.ln(4)
                pdf.set_font("Helvetica", "B", 12)
                pdf.cell(0, 8, safe(title), new_x="LMARGIN", new_y="NEXT")
                if df.empty:
                    pdf.set_font("Helvetica", "", 9)
                    pdf.cell(0, 6, "No data available.", new_x="LMARGIN", new_y="NEXT")
                    return
                pdf.set_font("Helvetica", "B", 8)
                for w, h in zip(col_widths, df.columns):
                    pdf.cell(w, 7, safe(str(h))[:28], border=1)
                pdf.ln()
                pdf.set_font("Helvetica", "", 8)
                for _, r in df.iterrows():
                    for w, v in zip(col_widths, r):
                        pdf.cell(w, 6, safe(str(v))[:30], border=1)
                    pdf.ln()

            def draw_bar_chart(title, labels, values, chart_width=180, bar_height=6, gap=2):
                """Hand-drawn horizontal bar chart -- no image rendering, no Chrome needed."""
                pdf.ln(4)
                pdf.set_font("Helvetica", "B", 12)
                pdf.cell(0, 8, safe(title), new_x="LMARGIN", new_y="NEXT")
                if not len(values):
                    pdf.set_font("Helvetica", "", 9)
                    pdf.cell(0, 6, "No data available.", new_x="LMARGIN", new_y="NEXT")
                    return
                max_val = max(values) if max(values) > 0 else 1
                label_width = 65
                bar_area = chart_width - label_width - 20
                x0 = pdf.get_x()
                for label, val in zip(labels, values):
                    y0 = pdf.get_y()
                    pdf.set_font("Helvetica", "", 7)
                    pdf.cell(label_width, bar_height, safe(str(label))[:38], new_x="LMARGIN", new_y="TOP")
                    bar_len = max(1, (val / max_val) * bar_area)
                    pdf.set_fill_color(139, 90, 60)
                    pdf.rect(x0 + label_width, y0, bar_len, bar_height, style="F")
                    pdf.set_xy(x0 + label_width + bar_len + 2, y0)
                    pdf.set_font("Helvetica", "", 7)
                    pdf.cell(20, bar_height, f"{val:,.0f}", new_x="LMARGIN", new_y="NEXT")
                    pdf.set_x(x0)
                    pdf.set_y(y0 + bar_height + gap)

            if not report_trend_df.empty:
                recent_trend = report_trend_df.tail(12)
                draw_bar_chart("Overall trend — actual kg, last 12 weeks",
                                recent_trend["period"].tolist(), recent_trend["kg"].tolist())

            if not report_bar_df.empty:
                bar_top = report_bar_df.sort_values("forecast_kg", ascending=False).head(12)
                draw_bar_chart("Forecast by segment (top 12)",
                                (bar_top["channel"] + " - " + bar_top["product"]).tolist(),
                                bar_top["forecast_kg"].tolist())

            draw_table("Translated forecast - kg and CAD",
                       report_dollar_df.assign(**{"Forecast (CAD)": report_dollar_df["Forecast (CAD)"].map(lambda x: f"${x:,.0f}")}) if not report_dollar_df.empty else report_dollar_df,
                       [45, 60, 30, 30])
            draw_table("Forecast accuracy overview", report_overview_df, [40, 50, 25, 25, 25, 20])
            pdf.add_page()
            draw_table("Forecast accuracy - MAPE and bias", report_accuracy_df, [45, 55, 25, 25, 30])

            pdf.ln(6)
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(107, 98, 88)
            pdf.multi_cell(0, 5, "ALERT = actuals off by 15%+ from forecast over the last 4 weeks. WATCH = 8-15%. "
                                 "MAPE = average error size regardless of direction. Bias: positive = "
                                 "under-forecasting, negative = over-forecasting. "
                                 "Generated automatically from 49th Parallel's demand planning app.")
            return bytes(pdf.output())

        def build_report_excel():
            from openpyxl.chart import BarChart, LineChart, Reference

            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Summary"
            ws1["A1"] = "Demand Planning Report"
            ws1["A1"].font = Font(bold=True, size=14)
            ws1["A2"] = f"Cycle {cycle} — generated {gen_time} — as of week of {latest_wk}"
            if cap_status_text:
                ws1["A4"] = cap_status_text
                ws1["A4"].font = Font(bold=True, color="B3432F" if cap_shortfall else "4A7A5C")

            def write_df(ws, df, start_row=1):
                for j, col in enumerate(df.columns, start=1):
                    c = ws.cell(row=start_row, column=j, value=col)
                    c.font = Font(bold=True, color="FFFFFF")
                    c.fill = PatternFill("solid", fgColor="1F4E78")
                for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
                    for j, val in enumerate(row, start=1):
                        ws.cell(row=i, column=j, value=val)
                for j, col in enumerate(df.columns, start=1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = max(14, len(str(col)) + 2)

            # trend chart -- native Excel line chart, built from real data written into the sheet
            ws_trend = wb.create_sheet("Overall Trend")
            trend_for_excel = report_trend_df.rename(columns={"period": "Week", "kg": "Total kg"})
            write_df(ws_trend, trend_for_excel)
            if len(trend_for_excel) > 1:
                chart1 = LineChart()
                chart1.title = "Overall trend — actual kg"
                chart1.y_axis.title = "Total kg"
                chart1.x_axis.title = "Week"
                data_ref = Reference(ws_trend, min_col=2, min_row=1, max_row=len(trend_for_excel) + 1)
                cats_ref = Reference(ws_trend, min_col=1, min_row=2, max_row=len(trend_for_excel) + 1)
                chart1.add_data(data_ref, titles_from_data=True)
                chart1.set_categories(cats_ref)
                chart1.width, chart1.height = 24, 10
                ws_trend.add_chart(chart1, "D2")

            ws2 = wb.create_sheet("Translated Forecast")
            if not report_dollar_df.empty:
                bar_data = report_dollar_df.sort_values("Forecast (kg)", ascending=False).copy()
                bar_data.insert(0, "Segment", bar_data["Channel"] + " - " + bar_data["Item"])
                write_df(ws2, bar_data)
                chart2 = BarChart()
                chart2.type = "bar"
                chart2.title = "Forecast by segment (kg)"
                data_ref2 = Reference(ws2, min_col=4, min_row=1, max_row=len(bar_data) + 1)
                cats_ref2 = Reference(ws2, min_col=1, min_row=2, max_row=len(bar_data) + 1)
                chart2.add_data(data_ref2, titles_from_data=True)
                chart2.set_categories(cats_ref2)
                chart2.width, chart2.height = 24, 12
                ws2.add_chart(chart2, "G2")

            ws3 = wb.create_sheet("Accuracy Overview")
            write_df(ws3, report_overview_df)

            ws4 = wb.create_sheet("Accuracy - MAPE and Bias")
            write_df(ws4, report_accuracy_df)

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        col1, col2, col3 = st.columns(3)
        with col1:
            st.download_button("Download report (HTML)", build_report_html(),
                                f"demand_report_{cycle}.html", mime="text/html")
        with col2:
            st.download_button("Download report (PDF)", build_report_pdf(),
                                f"demand_report_{cycle}.pdf", mime="application/pdf")
        with col3:
            st.download_button("Download report (Excel)", build_report_excel(),
                                f"demand_report_{cycle}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.caption("HTML opens in any browser and can be printed to PDF from there too. "
                   "PDF and Excel are generated directly, ready to attach or print for a meeting.")

# --- TAB 1: Upload ---
with tab_data:
    st.subheader("Upload raw sales records")
    st.caption(
        "One row per sale line: channel, customer (optional), product, size, kg sold, revenue, and date. "
        "Export this from Lightspeed/Acumatica and upload as CSV. Uploading new data automatically "
        "updates the forecast and checks last week's prediction."
    )
    uploaded = st.file_uploader("CSV or Excel file", type=["csv", "xlsx", "xls"])

    if uploaded is not None:
        if uploaded.name.lower().endswith((".xlsx", ".xls")):
            raw = pd.read_excel(uploaded)
        else:
            raw = pd.read_csv(uploaded)
        st.write("Preview:")
        st.dataframe(raw.head(), use_container_width=True)

        st.markdown("**Map your columns**")
        cols = list(raw.columns)
        col_defaults = load_upload_column_defaults()
        st.caption("Remembers your choices from last time — just confirm they still look right." if col_defaults else "")
        c1, c2, c3 = st.columns(3)
        with c1:
            date_col = st.selectbox("Date column", cols, index=default_index(cols, col_defaults.get("date_col")),
                                     help="Required now — the forecast is built from real dates.")
            channel_col = st.selectbox("Channel column", cols, index=default_index(cols, col_defaults.get("channel_col")))
            pt_options = ["(not available — derive it automatically)"] + cols
            product_type_col = st.selectbox(
                "Single vs Staple column", pt_options,
                index=default_index(pt_options, col_defaults.get("product_type_col")),
                help="Optional — lets you break the forecast down by Single vs Staple, and run each "
                     "with its own forecast (matching your earlier SARIMA coursework: staple ~6.4% "
                     "MAPE, single-origin ~2.9% MAPE, on their own history). Real Acumatica exports "
                     "usually don't have this column directly — pick 'derive it automatically' and "
                     "the app will work it out from the Product column instead.")
        with c2:
            cust_options = ["(not available)"] + cols
            customer_col = st.selectbox(
                "Customer column", cust_options, index=default_index(cust_options, col_defaults.get("customer_col")),
                help="Optional — many Acumatica exports don't include this.")
            product_col = st.selectbox("Product column", cols, index=default_index(cols, col_defaults.get("product_col")))
        with c3:
            size_col = st.selectbox("Size / package column", cols, index=default_index(cols, col_defaults.get("size_col")))
            revenue_col = st.selectbox("Revenue ($) column", cols, index=default_index(cols, col_defaults.get("revenue_col")))

        kg_mode_options = ["I have a direct KG column", "I have Units + Weight-per-unit (kg)"]
        kg_mode = st.radio("How is weight recorded?", kg_mode_options,
                            index=default_index(kg_mode_options, col_defaults.get("kg_mode")))
        if kg_mode == "I have a direct KG column":
            kg_col = st.selectbox("KG column", cols, index=default_index(cols, col_defaults.get("kg_col")))
            qty_options = ["(not available)"] + cols
            quantity_col = st.selectbox(
                "Quantity / bag count column (optional)", qty_options,
                index=default_index(qty_options, col_defaults.get("quantity_col")),
                help="How many units/bags each row represents — used to work out real kg-per-bag "
                     "rates, so forecasts can be converted into 'how many bags to order', not just kg.")
        else:
            units_col = st.selectbox("Units column", cols, index=default_index(cols, col_defaults.get("units_col")))
            weight_col = st.selectbox("Weight per unit (kg) column", cols, index=default_index(cols, col_defaults.get("weight_col")))
            quantity_col = units_col  # already exactly what's needed -- don't ask twice

        needs_auto_classify = product_type_col == "(not available — derive it automatically)"
        manual_overrides_by_item = {}
        classification_hint_col = None
        if needs_auto_classify:
            st.markdown("**Single vs Staple — classify by product, remembered across uploads**")
            st.caption(
                "Classification is keyed to the Product column above (e.g. an Inventory ID like "
                "'EE12') — that's what gets stored and needs a Staple/Single label going forward. "
                "But that field is often just a short code without useful words in it, so pick a "
                "different column below (like Item Class) that actually contains classification "
                "hints — this is only used to help detect the label, not stored as the product itself."
            )
            classification_hint_col = st.selectbox(
                "Which column has classification hints (e.g. Item Class)?",
                ["(use the Product column itself)"] + cols, key="classification_hint_col")
            hint_source = raw[product_col] if classification_hint_col == "(use the Product column itself)" \
                else raw[classification_hint_col]

            known_map = load_known_classifications()
            preview_products = raw[product_col].dropna().unique().tolist()
            audit_rows = []
            newly_learned = []
            for p in preview_products:
                if p in known_map:
                    audit_rows.append({"Product": p, "Classification": known_map[p], "Source": "Remembered from before"})
                else:
                    hint_val = raw.loc[raw[product_col] == p, classification_hint_col
                                        if classification_hint_col != "(use the Product column itself)" else product_col].iloc[0]
                    auto = classify_product_type_auto(hint_val)
                    if auto != "Unknown":
                        audit_rows.append({"Product": p, "Classification": auto, "Source": "Auto-detected (new)"})
                        newly_learned.append((p, auto))
                    else:
                        audit_rows.append({"Product": p, "Classification": "Unknown", "Source": "Needs manual input"})

            audit_df = pd.DataFrame(audit_rows)
            st.dataframe(audit_df, use_container_width=True, hide_index=True)
            unknown_items = audit_df[audit_df["Classification"] == "Unknown"]["Product"].tolist()

            if unknown_items:
                st.warning(f"{len(unknown_items)} product(s) couldn't be automatically classified — "
                           f"assign them once below, and this app will remember your answer for every "
                           f"future upload of the same product.")
                for item in unknown_items:
                    manual_overrides_by_item[item] = st.selectbox(
                        f"'{item}' is:", ["Staple", "Single"], key=f"manual_classify_{item}")
            else:
                st.success("Every product is classified — remembered from before, or auto-detected just now.")

        batch_name = st.text_input("Name this upload batch", value=f"upload_{datetime.now().strftime('%Y%m%d_%H%M')}")

        if st.button("Process and save", type="primary"):
            std = pd.DataFrame()
            std["record_date"] = raw[date_col].astype(str)
            std["channel"] = raw[channel_col].astype(str)
            std["customer"] = raw[customer_col].astype(str) if customer_col != "(not available)" else "(not tracked)"
            std["product"] = raw[product_col].astype(str)
            if needs_auto_classify:
                known_map = load_known_classifications()
                effective_hint_col = product_col if classification_hint_col == "(use the Product column itself)" else classification_hint_col

                def resolve_and_remember(p, hint_val):
                    p = str(p)
                    if p in known_map:
                        return known_map[p]
                    if p in manual_overrides_by_item:
                        save_classification(p, manual_overrides_by_item[p], "manual")
                        return manual_overrides_by_item[p]
                    auto = classify_product_type_auto(hint_val)
                    if auto != "Unknown":
                        save_classification(p, auto, "auto")
                        return auto
                    return "(not tracked)"

                std["product_type"] = [resolve_and_remember(p, h) for p, h in zip(raw[product_col], raw[effective_hint_col])]
            else:
                std["product_type"] = raw[product_type_col].astype(str)
            std["size_label"] = raw[size_col].astype(str)
            std["revenue"] = pd.to_numeric(raw[revenue_col], errors="coerce")
            if kg_mode == "I have a direct KG column":
                std["kg"] = pd.to_numeric(raw[kg_col], errors="coerce")
                std["quantity"] = pd.to_numeric(raw[quantity_col], errors="coerce") if quantity_col != "(not available)" else np.nan
            else:
                std["kg"] = pd.to_numeric(raw[units_col], errors="coerce") * pd.to_numeric(raw[weight_col], errors="coerce")
                std["quantity"] = pd.to_numeric(raw[quantity_col], errors="coerce")
            std = std.dropna(subset=["kg", "revenue"])
            std["upload_batch"] = batch_name
            std["uploaded_at"] = datetime.now().isoformat()
            insert_dataframe("sales_records", std, show_progress=True)

            mapping_to_remember = {
                "date_col": date_col, "channel_col": channel_col, "product_type_col": product_type_col,
                "customer_col": customer_col, "product_col": product_col, "size_col": size_col,
                "revenue_col": revenue_col, "kg_mode": kg_mode,
            }
            if kg_mode == "I have a direct KG column":
                mapping_to_remember.update({"kg_col": kg_col, "quantity_col": quantity_col})
            else:
                mapping_to_remember.update({"units_col": units_col, "weight_col": weight_col})
            save_upload_column_defaults(mapping_to_remember)

            st.success(f"Saved {len(std)} records from batch '{batch_name}'. Forecast will update below.")
            st.rerun()

    st.divider()
    st.caption(f"Total records in database: {len(sales_df)}")
    if has_data:
        batches = sales_df["upload_batch"].unique().tolist()
        st.write("Batches uploaded so far:", ", ".join(batches))

        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Delete a specific batch**")
            batch_to_delete = st.selectbox("Choose a batch", batches)
            if st.button("Delete this batch"):
                conn.execute("DELETE FROM sales_records WHERE upload_batch = ?", (batch_to_delete,))
                conn.commit()

                # real cleanup, not just a disclosed limitation -- figure out what the latest
                # actual week is NOW that this batch is gone, and remove any frozen forecasts
                # that were generated assuming data beyond that point existed. Without this,
                # a forecast frozen for the week right after the deleted batch stays sitting
                # in the database, and the app keeps showing/labeling around it as if that
                # batch were still there -- exactly the inconsistency this fixes.
                remaining_weekly = pd.read_sql("SELECT record_date, kg FROM sales_records", conn)
                if not remaining_weekly.empty:
                    remaining_weekly["record_date"] = pd.to_datetime(remaining_weekly["record_date"], errors="coerce")
                    remaining_weekly = remaining_weekly.dropna(subset=["record_date"])
                if remaining_weekly.empty:
                    conn.execute("DELETE FROM auto_forecasts")
                    removed_note = "No sales data remains, so all forecasts were cleared too."
                else:
                    remaining_weekly["week_start"] = (remaining_weekly["record_date"] -
                        pd.to_timedelta(remaining_weekly["record_date"].dt.weekday, unit="D")).dt.date.astype(str)
                    new_latest_week = sorted(remaining_weekly["week_start"].unique())[-1]
                    new_boundary = (pd.Timestamp(new_latest_week) + pd.Timedelta(days=7)).date().isoformat()
                    orphaned = pd.read_sql("SELECT id FROM auto_forecasts WHERE target_week > ?",
                                            conn, params=(new_boundary,))
                    if not orphaned.empty:
                        conn.execute("DELETE FROM auto_forecasts WHERE target_week > ?", (new_boundary,))
                        removed_note = f"Also removed {len(orphaned)} forecast(s) that were generated assuming this batch's data existed."
                    else:
                        removed_note = "No orphaned forecasts needed removing."
                conn.commit()

                st.warning(f"Deleted batch '{batch_to_delete}'. {removed_note}")
                st.rerun()
        with col2:
            st.markdown("**Or clear everything**")
            if st.button("Clear ALL uploaded records (careful)"):
                conn.execute("DELETE FROM sales_records")
                conn.execute("DELETE FROM auto_forecasts")
                conn.commit()
                st.warning("All sales records and forecasts cleared.")
                st.rerun()

# --- TAB 2: Computed rates ---
with tab_rates:
    st.subheader("Rates computed from your uploaded data")
    if not has_data:
        st.warning("No sales data uploaded yet — go to tab 1 first.")
    else:
        st.markdown("**Price per kg** — computed from your actual revenue and kg sold, channel x item x size, "
                     "last 120 days")
        display_price = price_df.copy()
        display_price["$ per kg"] = display_price["price_per_kg"].round(2)
        display_price["kg per $1 CAD"] = (1 / display_price["price_per_kg"].replace(0, np.nan)).round(4)
        if "price_min" in display_price.columns:
            display_price["customer_spread"] = display_price.apply(
                lambda r: f"${r['price_min']:.2f}–${r['price_max']:.2f}" if pd.notna(r.get("price_min")) else "n/a",
                axis=1)
            st.dataframe(display_price[["channel", "product", "size_label", "$ per kg", "kg per $1 CAD", "customer_spread"]]
                         .rename(columns={"customer_spread": "actual range by customer"}),
                         use_container_width=True)
            st.caption("**$ per kg** — how much revenue one kilo brings in (the more familiar framing for pricing). "
                       "**kg per $1 CAD** — the flip side: how many kilos $1 buys, useful when thinking in terms "
                       "of a budget (e.g. a $10,000 budget at 0.03 kg per $1 ≈ 300 kg). Both describe the exact "
                       "same rate, just read in whichever direction is more useful for what you're doing. "
                       "'Actual range by customer' shows how much the average is blending together — a wide "
                       "range means some customers pay meaningfully more or less than the average shown.")
        else:
            st.dataframe(price_df, use_container_width=True)
        st.download_button("Download price_per_kg.csv", price_df.to_csv(index=False), "price_per_kg.csv")

        st.markdown("**Price per kg by customer** — real customer-specific pricing, where there's enough "
                     "of that customer's own history to trust it")
        customer_price_df = compute_customer_price_per_kg(sales_df)
        if customer_price_df.empty:
            st.info("This data source doesn't include customer identity — customer-specific pricing isn't available.")
        else:
            display_cust_price = customer_price_df.copy()
            display_cust_price["Rate used"] = np.where(
                display_cust_price["confident"], "Customer-specific", "Channel average (not enough customer history)")
            display_cust_price = display_cust_price.rename(columns={
                "channel": "Channel", "customer": "Customer", "product": "Item",
                "customer_price_per_kg": "This customer's price/kg", "price_per_kg_used": "Price used in translation",
                "n_transactions": "Transactions"})
            st.dataframe(
                display_cust_price[["Channel", "Customer", "Item", "This customer's price/kg",
                                     "Price used in translation", "Transactions", "Rate used"]]
                .sort_values(["Channel", "Item", "Price used in translation"], ascending=[True, True, False]),
                use_container_width=True)
            n_confident = int(customer_price_df["confident"].sum())
            st.caption(f"{n_confident} of {len(customer_price_df)} customer/item combinations have enough "
                       f"transactions (3+) for a real customer-specific rate — the rest fall back to the "
                       f"channel average, since a price from 1-2 orders is noise, not a rate.")
            st.download_button("Download customer_price_per_kg.csv", customer_price_df.to_csv(index=False),
                                "customer_price_per_kg.csv")

        st.markdown("**Size mix %** (within each channel-product)")
        st.dataframe(size_mix_df, use_container_width=True)

        st.markdown("**Customer mix %** (within each channel-product, highest share first)")
        if customer_mix_df.empty or (customer_mix_df["customer"] == "(not tracked)").all():
            st.info("This data source doesn't include customer identity — customer-level breakdown isn't "
                     "available. Everything else still works fine without it.")
        else:
            st.dataframe(customer_mix_df, use_container_width=True)
            st.download_button("Download customer_mix.csv", customer_mix_df.to_csv(index=False), "customer_mix.csv")

# --- TAB 3: Forecast (auto) ---
with tab_forecast:
    st.subheader("Auto-generated forecast — no manual entry")
    st.caption(
        "This forecast isn't typed in — it's calculated from your own uploaded sales history. When there's "
        "enough history (8+ weeks), it uses ARIMA (tested against this data: ~13.5% average error, vs 15.2% "
        "for the simpler fallback method). With less history, it falls back to a trend method (median of the "
        "last 4 weeks, damped growth vs the prior 4) since ARIMA is unreliable on very short series. Plus "
        "anything logged in Pipeline / known events. It updates automatically every time you upload new data."
    )

    st.markdown("**Check and fix unstable historical forecasts**")
    st.caption(
        "Forecasts are frozen once generated, so a fix to the forecasting method only affects NEW forecasts "
        "going forward — it can't reach back and correct a number already stored. This checks every stored "
        "forecast against what a safe, bounded method would have predicted at the time, and corrects any "
        "that are wildly unstable (e.g. a spike caused by ARIMA overreacting to a short-lived blip)."
    )
    if st.button("Check and fix now", key="audit_forecasts_btn"):
        with st.spinner("Checking every stored forecast..."):
            checked, fixed, examples = audit_and_fix_historical_forecasts(weekly_actual)
        if fixed == 0:
            st.success(f"Checked {checked} stored forecasts — all within a sane range, nothing to fix.")
        else:
            st.warning(f"Checked {checked} stored forecasts, corrected {fixed} that were wildly unstable:")
            example_df = pd.DataFrame(examples, columns=["Channel", "Product", "Week", "Was", "Now"])
            st.dataframe(example_df, use_container_width=True, hide_index=True)
            st.info("Corrected — refresh the page or switch tabs to see the updated numbers everywhere.")

    st.divider()

    if forecast_by_cp.empty:
        st.warning("Not enough history yet — upload a few weeks of sales data in tab 1 first.")
    else:
        display = forecast_by_cp.copy()
        if not active_overrides.empty:
            display["Overridden"] = display.apply(
                lambda r: "Yes" if ((active_overrides["channel"] == r["channel"]) &
                                     (active_overrides["product"] == r["product"])).any() else "", axis=1)
        else:
            display["Overridden"] = ""
        display = display.rename(columns={"forecast_kg": "Forecast_kg (incl. pipeline + overrides)", "pipeline_kg": "Pipeline_adjustment_kg"})
        st.dataframe(display[["channel", "product", "target_week", "Pipeline_adjustment_kg",
                               "Forecast_kg (incl. pipeline + overrides)", "Overridden"]],
                     use_container_width=True)
        total_kg = forecast_by_cp["forecast_kg"].sum()
        st.metric("Total forecast kg (next unforecasted week, all channels/products)", f"{total_kg:,.0f} kg")

        if not translated.empty:
            st.markdown("**Broken down by size**")
            st.dataframe(translated, use_container_width=True)
            st.download_button("Download forecast_by_size.csv", translated.to_csv(index=False), "forecast_by_size.csv")

    st.divider()
    st.subheader("Manual override — replace the number directly")
    st.caption(
        "Pipeline events add a reason-backed adjustment on top of the auto forecast. This is different: "
        "it directly replaces the auto+pipeline number for a channel/item with whatever you type — for "
        "when you just know better than the trend, without a specific logged event to point to. "
        "An active override wins over everything else, and shows up in every table and KPI above and "
        "throughout the app (Dashboard, Ops capacity, translated $) — this is the one number everything else builds on."
    )

    current_target_week = None
    with st.form("override_form"):
        oc1, oc2 = st.columns(2)
        with oc1:
            ov_channel = st.selectbox("Channel", sorted(sales_df["channel"].unique().tolist())
                                       if has_data else ["Channel"], key="ov_channel")
            ov_product = st.selectbox("Item", sorted(sales_df["product"].unique().tolist())
                                       if has_data else ["Item"], key="ov_product")
        with oc2:
            ov_kg = st.number_input("Override forecast (kg)", min_value=0.0, step=10.0)
            ov_by = st.text_input("Your name")
        ov_period = st.radio(
            "How long should this apply?",
            ["One-time (just the current forecast period)", "Ongoing (until I turn it off)"],
            help="One-time automatically expires and reverts to the auto forecast once new data moves the "
                 "forecast to the next period. Ongoing keeps applying every time until you manually turn it off.")
        ov_note = st.text_area("Why are you overriding this? (kept for the record, doesn't gate the override)")

        if st.form_submit_button("Set override", type="primary"):
            if not live_forecast.empty:
                match_row = live_forecast[(live_forecast["channel"] == ov_channel) & (live_forecast["product"] == ov_product)]
                current_target_week = match_row.iloc[0]["target_week"] if not match_row.empty else None
            period_type = "One-time" if ov_period.startswith("One-time") else "Ongoing"
            conn.execute("UPDATE manual_overrides SET active = 0 WHERE channel = ? AND product = ?",
                         (ov_channel, ov_product))
            conn.execute("""INSERT INTO manual_overrides
                (timestamp, submitted_by, channel, product, override_kg, note, active, period_type, target_week)
                VALUES (?,?,?,?,?,?,1,?,?)""",
                (datetime.now().isoformat(), ov_by, ov_channel, ov_product, ov_kg, ov_note, period_type, current_target_week))
            conn.commit()
            st.success(f"Override set — {ov_channel} / {ov_product} now forecasts at {ov_kg:,.0f} kg "
                       f"({period_type.lower()}), everywhere in the app, replacing the auto+pipeline number.")
            st.rerun()

    active_overrides_display = pd.read_sql("SELECT * FROM manual_overrides WHERE active = 1 ORDER BY id DESC", conn)
    if active_overrides_display.empty:
        st.info("No active overrides — every forecast is currently coming from the auto method + pipeline events.")
    else:
        st.markdown("**Active overrides**")
        st.dataframe(active_overrides_display[["channel", "product", "override_kg", "period_type",
                                                 "submitted_by", "note", "timestamp"]],
                     use_container_width=True)
        turn_off_label = st.selectbox(
            "Turn off an override (reverts that channel/item back to the auto forecast)",
            (active_overrides_display["channel"] + " — " + active_overrides_display["product"]).tolist())
        if st.button("Turn off this override"):
            sel_ch, sel_pr = turn_off_label.split(" — ", 1)
            conn.execute("UPDATE manual_overrides SET active = 0 WHERE channel = ? AND product = ? AND active = 1",
                         (sel_ch, sel_pr))
            conn.commit()
            st.warning("Override turned off — reverting to the auto forecast.")
            st.rerun()

    # real history, not just what's currently active -- turning an override off updates its
    # status but never deletes the record, yet nowhere in the app previously showed that
    # history. This fixes a real gap: another team should be able to see what was overridden
    # in the past, by whom, and why, not just what's active right now.
    all_overrides_history = pd.read_sql("SELECT * FROM manual_overrides ORDER BY id DESC", conn)
    if not all_overrides_history.empty:
        with st.expander("Override history (active and turned-off)"):
            history_display = all_overrides_history.copy()
            history_display["status"] = history_display["active"].map({1: "Active", 0: "Turned off"})
            st.dataframe(history_display[["channel", "product", "override_kg", "period_type",
                                           "submitted_by", "note", "timestamp", "status"]],
                         use_container_width=True)

# --- TAB: Sales plan (S&OP) ---
with tab_salesplan:
    st.subheader("Sales plan — the top-down half of the S&OP bridge")
    st.caption(
        "This is Sales' own forward-looking plan — entered by month, channel, and item — translated "
        "into kg using the same price/kg rates used everywhere else in this app. It's compared against "
        "the app's own demand-sensing forecast below, so a gap between 'what Sales planned' and "
        "'what's actually happening' is visible early, not discovered at year-end."
    )

    plan_year = st.text_input("Plan year", value=str(date.today().year), key="plan_year")

    st.markdown("### Upload a plan (bulk)")
    plan_file = st.file_uploader("CSV or Excel", type=["csv", "xlsx", "xls"], key="plan_upload")
    if plan_file is not None:
        plan_raw = pd.read_excel(plan_file) if plan_file.name.lower().endswith((".xlsx", ".xls")) else pd.read_csv(plan_file)
        st.dataframe(plan_raw.head(), use_container_width=True)
        pcols = list(plan_raw.columns)
        pc1, pc2, pc3 = st.columns(3)
        with pc1:
            plan_channel_col = st.selectbox("Channel column", pcols, key="plan_channel_col")
            plan_month_col = st.selectbox("Month column", pcols, key="plan_month_col")
        with pc2:
            plan_product_col = st.selectbox("Item column", pcols, key="plan_product_col")
            plan_amount_col = st.selectbox("Planned amount column", pcols, key="plan_amount_col")
        with pc3:
            plan_amount_type = st.radio("This amount is in", ["Dollars ($)", "Kilograms (kg)"], key="plan_amount_type")

        if st.button("Save this plan", type="primary"):
            std_plan = pd.DataFrame()
            std_plan["channel"] = plan_raw[plan_channel_col].astype(str)
            std_plan["product"] = plan_raw[plan_product_col].astype(str)
            std_plan["month"] = plan_raw[plan_month_col].astype(str)
            if plan_amount_type == "Dollars ($)":
                std_plan["planned_dollars"] = pd.to_numeric(plan_raw[plan_amount_col], errors="coerce")
                std_plan["planned_kg"] = np.nan
            else:
                std_plan["planned_kg"] = pd.to_numeric(plan_raw[plan_amount_col], errors="coerce")
                std_plan["planned_dollars"] = np.nan
            std_plan = std_plan.dropna(subset=["channel", "product", "month"])

            if not price_df.empty:
                rate_lookup = price_df.groupby(["channel", "product"], as_index=False)["price_per_kg"].mean()
                std_plan = std_plan.merge(rate_lookup, on=["channel", "product"], how="left")
                std_plan["planned_kg"] = np.where(
                    std_plan["planned_kg"].isna() & std_plan["planned_dollars"].notna() & std_plan["price_per_kg"].notna(),
                    std_plan["planned_dollars"] / std_plan["price_per_kg"], std_plan["planned_kg"])
                std_plan = std_plan.drop(columns=["price_per_kg"])

            std_plan["plan_year"] = plan_year
            std_plan["updated_at"] = datetime.now().isoformat()
            std_plan["updated_by"] = ""
            std_plan["note"] = ""
            insert_dataframe("sales_plan", std_plan)
            st.success(f"Saved {len(std_plan)} plan rows for {plan_year}.")
            st.rerun()

    st.divider()
    st.markdown("### Edit the current plan")
    st.caption("Add, edit, or delete rows directly — this is how Sales keeps the plan updated over time.")
    existing_plan = pd.read_sql("SELECT * FROM sales_plan WHERE plan_year = ? ORDER BY month, channel, product",
                                 conn, params=(plan_year,))
    edit_base = existing_plan[["channel", "product", "month", "planned_dollars", "planned_kg", "note"]] \
        if not existing_plan.empty else pd.DataFrame(columns=["channel", "product", "month", "planned_dollars", "planned_kg", "note"])

    edited = st.data_editor(edit_base, num_rows="dynamic", use_container_width=True, key="plan_editor")
    updated_by = st.text_input("Your name (for the record)", key="plan_updated_by_input")

    if st.button("Save changes to plan", type="primary"):
        edited_clean = edited.dropna(subset=["channel", "product", "month"], how="any").copy()
        if not edited_clean.empty and not price_df.empty:
            rate_lookup = price_df.groupby(["channel", "product"], as_index=False)["price_per_kg"].mean()
            edited_clean = edited_clean.merge(rate_lookup, on=["channel", "product"], how="left")
            edited_clean["planned_kg"] = np.where(
                edited_clean["planned_kg"].isna() & edited_clean["planned_dollars"].notna() & edited_clean["price_per_kg"].notna(),
                edited_clean["planned_dollars"] / edited_clean["price_per_kg"], edited_clean["planned_kg"])
            edited_clean = edited_clean.drop(columns=["price_per_kg"])
        edited_clean["plan_year"] = plan_year
        edited_clean["updated_at"] = datetime.now().isoformat()
        edited_clean["updated_by"] = updated_by

        conn.execute("DELETE FROM sales_plan WHERE plan_year = ?", (plan_year,))
        conn.commit()
        if not edited_clean.empty:
            insert_dataframe("sales_plan", edited_clean)
        st.success(f"Plan for {plan_year} updated ({len(edited_clean)} rows).")
        st.rerun()

    st.divider()
    st.markdown("## Reconciliation — plan vs. reality")
    st.caption(
        "For months that already happened, this compares the plan to real actuals. For months still "
        "ahead, it compares the plan to the app's own demand-sensing projection — which gets less "
        "certain the further out it looks, unlike the plan (which usually assumes similar confidence "
        "across the whole year). Treat far-future gaps as a rough signal, not a precise miss."
    )

    plan_monthly = existing_plan.groupby("month", as_index=False)["planned_kg"].sum() if not existing_plan.empty else pd.DataFrame()

    if plan_monthly.empty:
        st.info("Enter a plan above to see the reconciliation view.")
    elif not has_data:
        st.info("Upload sales history in tab 1 to compare the plan against.")
    else:
        actuals_monthly = sales_df.copy()
        actuals_monthly["record_date"] = pd.to_datetime(actuals_monthly["record_date"], errors="coerce")
        actuals_monthly["month"] = actuals_monthly["record_date"].dt.to_period("M").astype(str)
        actuals_monthly = actuals_monthly.groupby("month", as_index=False)["kg"].sum().rename(columns={"kg": "actual_kg"})

        recon = plan_monthly.merge(actuals_monthly, on="month", how="left").sort_values("month")
        missing_months = recon[recon["actual_kg"].isna()]["month"].tolist()

        recon["demand_sensing_kg"] = np.nan
        if missing_months:
            company_monthly = sales_df.copy()
            company_monthly["record_date"] = pd.to_datetime(company_monthly["record_date"], errors="coerce")
            company_monthly["month"] = company_monthly["record_date"].dt.to_period("M").astype(str)
            company_monthly_agg = company_monthly.groupby("month", as_index=False)["kg"].sum().sort_values("month")
            if len(company_monthly_agg) >= 2:
                with st.spinner("Computing demand-sensing projection for remaining months..."):
                    proj_recon = project_forward_with_range(company_monthly_agg["kg"].tolist(), None,
                                                              n_periods=min(len(missing_months) + 3, 12))
                last_month = pd.Period(company_monthly_agg["month"].iloc[-1], freq="M")
                proj_months = [(last_month + i + 1).strftime("%Y-%m") for i in range(len(proj_recon))]
                proj_lookup = dict(zip(proj_months, proj_recon["forecast_kg"]))
                recon["demand_sensing_kg"] = recon["month"].map(proj_lookup)

        recon["Status"] = np.where(recon["actual_kg"].notna(), "Actual", "Forecast (projected)")
        recon["Reality (kg)"] = recon["actual_kg"].fillna(recon["demand_sensing_kg"])
        recon["Gap vs plan (kg)"] = recon["Reality (kg)"] - recon["planned_kg"]
        recon["Gap %"] = np.where(recon["planned_kg"] != 0, recon["Gap vs plan (kg)"] / recon["planned_kg"] * 100, np.nan)

        display_recon = recon.rename(columns={"month": "Month", "planned_kg": "Plan (kg)"})[
            ["Month", "Plan (kg)", "Reality (kg)", "Gap vs plan (kg)", "Gap %", "Status"]].round(1)
        st.dataframe(display_recon, use_container_width=True, hide_index=True)

        fig_recon = go.Figure()
        fig_recon.add_trace(go.Scatter(x=recon["month"], y=recon["planned_kg"], mode="lines+markers",
                                        name="Sales plan", line=dict(color="rgb(217,119,6)", width=2)))
        fig_recon.add_trace(go.Scatter(x=recon["month"], y=recon["Reality (kg)"], mode="lines+markers",
                                        name="Actual / demand-sensing", line=dict(color="rgb(31,119,180)", width=2)))
        fig_recon.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10), plot_bgcolor="white",
                                 yaxis_title="kg", hovermode="x unified")
        st.plotly_chart(fig_recon, use_container_width=True)

# --- TAB: Pipeline / known events ---
with tab_pipeline:
    st.subheader("Log what's happening right now — before it shows up in sales data")
    st.caption(
        "Signed a new contract? Lost an account? Know volume is about to change? Log it here. "
        "A trend forecast can't see a deal that hasn't shipped yet — but you already know about it. "
        "Once logged, it adjusts the forecast automatically, on top of the auto-generated baseline."
    )

    with st.form("pipeline_form"):
        c1, c2 = st.columns(2)
        with c1:
            event_type = st.selectbox("What happened", [
                "New contract signed", "Account lost / churned", "Expected volume change", "Other"])
            customer = st.text_input("Customer / account name")
            channel = st.selectbox("Channel", sorted(sales_df["channel"].unique().tolist())
                                    if has_data else ["Wholesale", "Retail"])
            product = st.selectbox("Product", sorted(sales_df["product"].unique().tolist())
                                    if has_data else ["Product"])
        with c2:
            expected_kg = st.number_input(
                "Expected kg impact per month", step=1.0,
                help="Positive for new/growing business, negative for a lost account.")
            starting_cycle = st.text_input("Starting cycle", value=cycle,
                                            help="Which cycle this first applies to, e.g. 2026-08")
            ongoing = st.checkbox("Ongoing (keeps applying to future cycles)", value=True,
                                   help="Uncheck for a one-time event that only affects this one cycle.")
            submitted_by = st.text_input("Logged by")
        note = st.text_area("Note (context for whoever reads this later)")

        if st.form_submit_button("Log this event", type="primary"):
            conn.execute("""INSERT INTO pipeline_events
                (timestamp, submitted_by, event_type, customer, channel, product,
                 expected_kg_per_month, starting_cycle, ongoing, note)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (datetime.now().isoformat(), submitted_by, event_type, customer, channel, product,
                 expected_kg, starting_cycle, int(ongoing), note))
            conn.commit()
            st.success("Logged — the forecast now includes this.")
            st.rerun()

    st.divider()
    if all_events.empty:
        st.info("No events logged yet.")
    else:
        st.markdown(f"**Currently applying to cycle {cycle}:**")
        if applicable.empty:
            st.caption("None of the logged events apply to this cycle yet.")
        else:
            st.dataframe(
                applicable[["event_type", "customer", "channel", "product",
                            "expected_kg_per_month", "starting_cycle", "ongoing", "note"]],
                use_container_width=True)
            total_pipeline_kg = applicable["expected_kg_per_month"].sum()
            st.metric("Total pipeline adjustment this cycle", f"{total_pipeline_kg:+,.0f} kg")

            st.markdown("**Is this actually changing the forecast? Here's the before/after.**")
            if not live_forecast.empty:
                impact = live_forecast[["channel", "product", "forecast_kg"]].rename(
                    columns={"forecast_kg": "Baseline forecast (kg) — trend only, no events"})
                impact = impact.merge(pipeline_by_cp, on=["channel", "product"], how="inner")
                if impact.empty:
                    st.warning("None of the logged events match a channel/product the trend model currently "
                               "has a forecast for — double check the Channel and Product you selected when "
                               "logging the event match what's actually in your uploaded sales data.")
                else:
                    impact["Adjusted forecast (kg) — with events"] = (
                        impact["Baseline forecast (kg) — trend only, no events"] + impact["pipeline_kg"])
                    impact = impact.rename(columns={"channel": "Channel", "product": "Item",
                                                     "pipeline_kg": "Event adjustment (kg)"})
                    st.dataframe(impact[["Channel", "Item", "Baseline forecast (kg) — trend only, no events",
                                          "Event adjustment (kg)", "Adjusted forecast (kg) — with events"]],
                                 use_container_width=True)
                    st.caption("If the baseline and adjusted columns are identical, the event isn't actually "
                               "reaching this channel/product — check the Channel/Product spelling matches "
                               "your sales data exactly.")
            else:
                st.info("No baseline forecast to compare against yet — upload more sales history first.")

        st.markdown("**Delete an event** (e.g. a contract that got cancelled)")
        event_labels = all_events.apply(
            lambda r: f"#{r['id']} — {r['event_type']} — {r['customer']} — {r['channel']}/{r['product']} "
                      f"({r['expected_kg_per_month']:+.0f} kg/mo, starts {r['starting_cycle']})", axis=1)
        event_map = dict(zip(event_labels, all_events["id"]))
        to_delete_label = st.selectbox("Choose an event", event_labels.tolist())
        if st.button("Delete this event"):
            conn.execute("DELETE FROM pipeline_events WHERE id = ?", (int(event_map[to_delete_label]),))
            conn.commit()
            st.warning("Deleted — the forecast updates immediately.")
            st.rerun()

        with st.expander("All logged events (all cycles)"):
            st.dataframe(all_events, use_container_width=True)

# --- TAB: Ops capacity check ---
with tab_ops:
    st.subheader("Operations: enter capacity and check against the plan")
    if forecast_by_cp.empty:
        st.warning("No forecast yet — upload sales data in tab 1 first.")
    else:
        existing_cap = pd.read_sql("SELECT * FROM ops_capacity WHERE cycle_label = ? ORDER BY id DESC LIMIT 1",
                                    conn, params=(cycle,))
        existing_cap = existing_cap.iloc[0] if not existing_cap.empty else None
        with st.form("ops_form"):
            ops_name = st.text_input("Your name", value=existing_cap["submitted_by"] if existing_cap is not None else "")
            monthly_capacity = st.number_input(
                "Monthly production capacity (kg)",
                value=float(existing_cap["monthly_capacity_kg"]) if existing_cap is not None else 4000.0)
            ops_notes = st.text_area("Notes", value=existing_cap["notes"] if existing_cap is not None else "")
            if st.form_submit_button("Save capacity", type="primary"):
                conn.execute("""INSERT INTO ops_capacity (timestamp, submitted_by, cycle_label, monthly_capacity_kg, notes)
                    VALUES (?,?,?,?,?)""", (datetime.now().isoformat(), ops_name, cycle, monthly_capacity, ops_notes))
                conn.commit()
                st.success("Capacity saved.")
                st.rerun()

        cap_row = pd.read_sql("SELECT * FROM ops_capacity WHERE cycle_label = ? ORDER BY id DESC LIMIT 1",
                               conn, params=(cycle,))
        if not cap_row.empty:
            cap = cap_row.iloc[0]["monthly_capacity_kg"]
            weekly_planned = forecast_by_cp["forecast_kg"].sum()
            monthly_planned = weekly_planned * 4.345
            status = "SHORTFALL" if cap < monthly_planned else "OK"
            st.metric("Monthly-equivalent planned kg vs capacity",
                       f"{monthly_planned:,.0f} kg planned / {cap:,.0f} kg capacity", status)
            if status == "SHORTFALL":
                st.error(f"Capacity shortfall of {monthly_planned - cap:,.0f} kg/month — flag for sign-off discussion.")
            else:
                st.success("Capacity covers this plan.")

# --- TAB 5: Sign-off ---
with tab_signoff:
    st.subheader("Consensus sign-off")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Sales sign-off**")
        sales_signer = st.text_input("Sales rep name", key="sales_signer")
        if st.button("Sign off — Sales"):
            conn.execute("INSERT INTO signoffs (timestamp, cycle_label, role, name) VALUES (?,?,?,?)",
                         (datetime.now().isoformat(), cycle, "Sales", sales_signer))
            conn.commit()
            st.success("Sales sign-off recorded.")
    with col2:
        st.markdown("**Ops sign-off**")
        ops_signer = st.text_input("Ops rep name", key="ops_signer")
        if st.button("Sign off — Ops"):
            conn.execute("INSERT INTO signoffs (timestamp, cycle_label, role, name) VALUES (?,?,?,?)",
                         (datetime.now().isoformat(), cycle, "Ops", ops_signer))
            conn.commit()
            st.success("Ops sign-off recorded.")

    signoffs = pd.read_sql("SELECT * FROM signoffs WHERE cycle_label = ? ORDER BY id DESC", conn, params=(cycle,))
    has_sales = (signoffs["role"] == "Sales").any() if not signoffs.empty else False
    has_ops = (signoffs["role"] == "Ops").any() if not signoffs.empty else False
    if has_sales and has_ops:
        st.success(f"Cycle {cycle} APPROVED by both sides.")
    else:
        st.info(f"Pending: {'Sales OK' if has_sales else 'Sales -'} / {'Ops OK' if has_ops else 'Ops -'}")
    st.dataframe(signoffs, use_container_width=True)

# --- TAB: Ask AI ---
with tab_ai:
    st.subheader("Ask AI about the current forecast")
    st.caption(
        "Ask questions grounded in this cycle's real data — the current forecast, capacity status, "
        "pipeline events, and recent accuracy. Costs a fraction of a cent per question."
    )

    api_key_set = hasattr(st, "secrets") and bool(st.secrets.get("ANTHROPIC_API_KEY"))
    if not api_key_set:
        st.warning("No API key set up yet.")
        with st.expander("How to set this up (one-time, ~2 minutes)"):
            st.markdown("""
1. Get an API key from **console.anthropic.com** (separate from a claude.ai login — this is API access, billed per use).
2. In Streamlit Community Cloud, open this app's settings → **Secrets**.
3. Add:
```
ANTHROPIC_API_KEY = "sk-ant-...your-key-here..."
```
4. Save. The app picks it up automatically, no redeploy needed.

The key is never visible in the code or GitHub repo — Streamlit's secrets storage keeps it separate.
            """)
    else:
        def build_context_summary():
            parts = [f"Planning cycle: {cycle}"]
            if not forecast_by_cp.empty:
                parts.append("Current forecast by channel/item (kg):\n" +
                              forecast_by_cp.to_string(index=False))
            if not dollar_by_cp.empty:
                parts.append("Translated forecast value (kg and CAD):\n" +
                              dollar_by_cp.to_string(index=False))
            cap_row_ai = pd.read_sql("SELECT * FROM ops_capacity WHERE cycle_label = ? ORDER BY id DESC LIMIT 1",
                                      conn, params=(cycle,))
            if not cap_row_ai.empty:
                parts.append(f"Ops capacity this cycle: {cap_row_ai.iloc[0]['monthly_capacity_kg']:,.0f} kg/month")
            if not applicable.empty:
                parts.append("Active pipeline events this cycle:\n" +
                              applicable[["event_type", "customer", "channel", "product",
                                          "expected_kg_per_month"]].to_string(index=False))
            if not backtest_df.empty:
                recent_summary = []
                for key, grp in backtest_df.groupby(["channel", "product"]):
                    bias = grp.sort_values("week_start")["variance_pct"].tail(4).mean()
                    if pd.notna(bias):
                        recent_summary.append(f"{key[0]} / {key[1]}: recent bias {bias*100:+.0f}%")
                parts.append("Recent forecast accuracy by segment:\n" + "\n".join(recent_summary[:30]))
            return "\n\n".join(parts)

        question = st.text_area("Your question", placeholder="e.g. Which segments should I be worried about this week, and why?")
        if st.button("Ask", type="primary") and question:
            with st.spinner("Asking..."):
                context = build_context_summary()
                prompt = (f"You are helping analyze a coffee roaster's demand planning data. "
                          f"Here is the current state:\n\n{context}\n\n"
                          f"Question: {question}\n\n"
                          f"Answer directly and specifically using the numbers above. If the data doesn't "
                          f"contain what's needed to answer, say so rather than guessing.")
                answer, err = call_claude(prompt)
                if err:
                    st.error(err)
                else:
                    st.markdown(answer)

    st.divider()
    st.subheader("Suggest a change to this app")
    st.caption(
        "Describe a change you'd like — this drafts a suggested code edit for you to review and apply "
        "yourself. Nothing is changed automatically; this never touches the live app or your GitHub repo."
    )

    if not api_key_set:
        st.info("Set up the API key above first.")
    else:
        change_request = st.text_area(
            "What would you like changed?",
            placeholder="e.g. Change the ALERT threshold from 15% to 20%, or add a filter for size on the Dashboard.")
        if st.button("Draft a suggestion") and change_request:
            with st.spinner("Reading the app and drafting a suggestion..."):
                try:
                    with open(__file__, "r") as f:
                        current_source = f.read()
                except Exception as e:
                    current_source = None
                    st.error(f"Couldn't read the app's own source: {e}")

                if current_source:
                    prompt = (
                        "You are helping a non-expert Streamlit developer modify their own app. "
                        "Below is the FULL current source code of their app.py. A user wants a specific change. "
                        "Suggest the exact, minimal code edit needed: show the exact snippet to find, and what "
                        "to replace it with. Keep the diff as small as possible. Briefly explain what it does "
                        "and flag anything risky (e.g. if it affects the database schema or other tabs). "
                        "Do not suggest unrelated changes.\n\n"
                        f"=== CURRENT app.py ===\n{current_source}\n=== END app.py ===\n\n"
                        f"Requested change: {change_request}"
                    )
                    answer, err = call_claude(prompt, max_tokens=2000)
                    if err:
                        st.error(err)
                    else:
                        st.markdown(answer)
                        st.warning("This is a suggestion only — copy the relevant part into your own editor, "
                                   "test it, then upload to GitHub to redeploy. Nothing here has been applied.")

# --- TAB 6: History ---
with tab_history:
    st.subheader("Backup / export everything")
    st.markdown("**Sales records**")
    st.dataframe(sales_df, use_container_width=True)
    st.download_button("Download sales_records.csv", sales_df.to_csv(index=False), "sales_records.csv")
    st.markdown("**Auto-generated forecasts (frozen predictions, for accuracy tracking)**")
    af = pd.read_sql("SELECT * FROM auto_forecasts ORDER BY id DESC", conn)
    st.dataframe(af, use_container_width=True)
    st.download_button("Download auto_forecasts.csv", af.to_csv(index=False), "auto_forecasts.csv")
    st.markdown("**Pipeline events**")
    st.dataframe(all_events, use_container_width=True)
    st.markdown("**Sign-offs**")
    st.dataframe(pd.read_sql("SELECT * FROM signoffs ORDER BY id DESC", conn), use_container_width=True)
    st.caption("Download buttons above back up all data as CSV — recommended periodically, "
               "since free-tier hosting can reset local storage on redeploy.")
